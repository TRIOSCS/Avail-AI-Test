"""Tests for the FRU matrix ingest command (app/management/ingest_fru_matrix.py).

What: Builds a miniature FRU_PN_TRAY workbook in-test with openpyxl covering ALL ten
      mapped sheets (Main, Qlot, Gabor, CZ, CDC, Lenovo-HDD, Lenovo FRU-PN, LVN VPD
      Mapping, Series, NSeries) plus a known-skipped sheet, exercising the verified
      hygiene cases — \xa0 non-breaking spaces, "N/ A"/"#N/A"/PENDIENTE sentinels,
      comma multi-values, carrier parentheticals, prose cells (FRU and related),
      self-edges, SAP zero-padded Lenovo FRUs, NSeries forward-fill — and asserts
      the parsed links plus the unparsed-cell accounting. Also covers the sheet
      coverage guard (missing mapped sheet fatal, unexpected sheet blocks --apply),
      context-column length caps, and the --apply upsert (insert new / refresh
      attributes / additive-only / rel_kind validation / no duplicates).
Called by: pytest
Depends on: openpyxl, app/management/ingest_fru_matrix.py, app.models.FruLink
"""

from datetime import datetime

import openpyxl
import pytest

from app.constants import CDC_PENDING, FruLinkKind
from app.management.ingest_fru_matrix import (
    PARSERS,
    ParsedLink,
    _clean,
    _Emitter,
    _lenovo_fru,
    _parse_carrier,
    _pn,
    _split_pns,
    main,
    parse_workbook,
    upsert_links,
)
from app.models import FruLink


@pytest.fixture()
def mini_workbook(tmp_path):
    """A miniature FRU matrix workbook with the real sheet names/layouts (all 10)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    main_ws = wb.create_sheet("Main")
    main_ws.append(
        [
            "FRU",
            "Part Number",
            "Option",
            "Opt-PN",
            "Model",
            "Assembly",
            "Additional Sourcing Numbers",
            "Manufactured by",
            "Carrier",
            "Alternate Carrier",
            "Machine",
            "Feature Code",
            "Comment",
            "Series",
        ]
    )
    # Plain row: 11S + model + carrier with parenthetical.
    main_ws.append(
        [
            "00AJ001",
            "68Y7789",
            "00AJ008",
            None,
            "SSDSC2BB120G4I",
            None,
            "40K1107, 42C0432",
            "Intel",
            "41Y0708 (w/interposer 41Y0709)",
            None,
            "x3650",
            "A2U4",
            "ThinkServer",
            "xSeries",
        ]
    )
    # Hygiene row: nbsp in PN, sentinel manufacturer, prose sourcing cell, slash carrier.
    main_ws.append(
        [
            "00AJ002",
            "\xa0 68Y7790 ",
            "N/A",
            "#N/A",
            "WD4000FYYZ-88UL1B0",
            None,
            "Approved sub is 81Y9727 w/ same 11S PN AND FW SN03, but should be acceptable.",
            "PENDIENTE",
            "SM10G01157 / 00FC544 (Blue hot-swap)",
            "N/ A",
            None,
            None,
            None,
            "pSeries",
        ]
    )
    # Row with no usable FRU (prose) and a fully blank padding row.
    main_ws.append(
        ["59Y5341 → rework in CZ", "12345AB", None, None, None, None, None, None, None, None, None, None, None, None]
    )
    main_ws.append([None] * 14)
    # Self-edge row: Part Number repeats the FRU → must NOT emit an ibm_11s link.
    main_ws.append(["00AJ003", "00AJ003", None, None, None, None, None, None, None, None, None, None, None, None])

    qlot = wb.create_sheet("Qlot as of 6.2025")
    qlot.append(["FRU/DRIVE", "FRU", "Drivepn", "Type", "IW CSP qual", "comment"])
    qlot.append(["00Y245900Y2430", "00Y2459", "00Y2430", "V3700", "qlot approved", "drop-in"])

    gabor = wb.create_sheet("Gabor 11.13.25")
    gabor.append(
        [
            "FRU/DRIVE",
            "FRU",
            "Drivepn",
            "Type",
            "IW CSP qual",
            "Qlot approved date",
            "Brand",
            "Bare drive description",
            "OEM",
            "FRU Description",
            "Tray",
            "Drive Model",
            None,
            "CDC",
        ]
    )
    gabor.append(
        [
            "03GU64900VN562",
            "03GU649",
            "00VN562",
            "Storwize",
            "qlot approved",
            datetime(2024, 3, 14),
            "V5000 Gen2",
            "WD Paris C Non SED 18TB",
            "WDC",
            "18TB 3.5 HDD 7.2K SAS",
            "01LJ138",
            "WUH721818AL4200",
            None,
            "x",
        ]
    )
    # Drive Model #N/A → no mfg_model link.
    gabor.append(
        [
            "01AC60200VN225",
            "01AC602",
            "00VN225",
            "Storwize",
            "qlot approved",
            None,
            None,
            "SSD, Micron, S650DC",
            "Micron",
            "1.6TB 2.5 SSD SAS",
            "45W8687",
            "#N/A",
            None,
            None,
        ]
    )

    cz = wb.create_sheet("CZ ONLY Testing")
    cz.append(["FRU/DRIVE", "FRU", "Drivepn", "Type", "qual", "date", "OEM", "Bare desc", "FRU desc", "Tray", "Model"])
    cz.append(
        [
            "01EJ73901EJ710",
            "01EJ739",
            "01EJ710",
            "V5000",
            "qlot approved - Only EMEA",
            datetime(2024, 5, 1),
            "Seagate",
            "Bare 8TB NL-SAS",
            "8TB 3.5 HDD",
            "01EJ900",
            "ST8000NM0185",
        ]
    )

    cdc = wb.create_sheet("CDC NOT yet AP Article Required")
    cdc.append(["FRU/DRIVE", "FRU", "Drivepn", "Model", "BRAND", "CDC Platform"])
    cdc.append(["01GV87001GV840", "01GV870", "01GV840", "10TB 7.2K NL SAS drive", "FlashSystem", "CDC4"])

    lhdd = wb.create_sheet("Lenovo-HDD")
    lhdd.append(
        [
            "FRU",
            "Part Number",
            "Lenovo PN",
            "Opt-PN",
            "Model",
            "x5",
            "x6",
            "Manufactured by",
            "Description",
            "Carrier",
            "Machine",
            "FW",
            "x12",
            "x13",
            "Series",
        ]
    )
    lhdd.append(
        [
            "00YK010",
            "00YK011",
            "SSS7A12345",
            None,
            "ST600MM0088",
            None,
            None,
            "Seagate",
            "600GB 10K SAS",
            "00E7600 (Gen3 tray)",
            "SR650",
            "LK69",
            None,
            None,
            "ThinkSystem",
        ]
    )

    lfp = wb.create_sheet("Lenovo FRU-PN")
    lfp.append(["BOM Type", "SBB Change Type", "SBB", "FRU", "Qty", "PPN", "Last Modified Time"])
    lfp.append(["FRU-PPN", "", "", "0000000NV340_E00", "", "ESG0017964", ""])  # padded + suffix
    lfp.append(["FRU-PPN", "", "", "SB17B49754", "", "0000049Y4746", ""])  # 10-char FRU, padded PPN
    lfp.append(["FRU-PPN", "", "", "", "", "ESG999", ""])  # blank FRU → skipped
    lfp.append(["FRU-PPN", "", "", "see SBB above", "", "ESG0017965", ""])  # prose FRU → skipped
    lfp.append(["FRU-PPN", "", "", "0000000NV341_E00", "", "no PPN - use SBB", ""])  # prose PPN → unparsed

    lvn = wb.create_sheet("LVN VPD Mapping")
    lvn.append(
        [
            "Brand",
            "x1",
            "x2",
            "x3",
            "Option P/N",
            "FRU P/N",
            "Option P/N 2",
            "MFG Model",
            "Make-to-Label P/N",
            "Tray P/N",
            "Description",
        ]
    )
    # Platform brand → note, not manufacturer; options from BOTH option columns.
    lvn.append(
        [
            "ThinkSystem",
            None,
            None,
            None,
            "4XB7A14099",
            "02JG542",
            "4XB7A14100",
            "MZILT3T8HBLS",
            "SSS7A23456",
            "01PG596",
            "3.84TB SSD",
        ]
    )
    # Non-platform brand → manufacturer.
    lvn.append(["Samsung", None, None, None, None, "02JG543", None, "MZILT7T6HALA", None, None, "7.68TB SSD"])

    series = wb.create_sheet("Series")
    series.append(
        [
            "fru",
            "partno",
            "idmodelo",
            "Manufacturer",
            "Capacity",
            "Size",
            "rpm x1000",
            "Pins",
            "partnosg",
            "bracket",
            "bracket_alt",
            "board",
            "board_alt",
            "screws",
            "wty",
            "cdc",
            "Alternate CDC-FRU",
            "Brand",
        ]
    )
    series.append(
        [
            "00P1517",
            "08K0263",
            "IC35L018UCDY10-0",
            "HITACHI/IBM",
            9,
            3.5,
            10,
            80,
            None,
            "U3X",
            "53P5457",
            "39J3525, 53P5456",
            None,
            "05J7985",
            60,
            0,
            None,
            "pSeries",
        ]
    )
    series.append(
        [
            "00P1518",
            "71P7530",
            "PENDIENTE",
            "SEAGATE",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "iSeries",
        ]
    )

    ns = wb.create_sheet("NSeries(NetApp)")
    ns.append(
        [
            "FRU",
            "Description",
            "3X5",
            "Shuttle Type",
            "Shuttle/",
            "Dongle P/N",
            "Dongle Screw P/N",
            "Seagate MFG P/N",
            "SG ",
            "Hitachi MFG P/N",
            "HIT FW",
            "Western Digital",
            "WD",
        ]
    )
    ns.append([None, None, None, None, "Screw P/N", None, None, None, "FW", None, None, None, "FW"])
    ns.append(
        [
            "45E1427",
            "HDD,FRU,1TB,7.2K RPM,SATA",
            "108-00183",
            "DS14MK2ATA",
            "X298A-R5, SP-298A-R5",
            "65695-02",
            "83315-xx",
            None,
            None,
            "0A35002 ",
            "A90A",
            "WD1002FBYS-05A6B0",
            "NA01",
        ]
    )
    # Continuation row: blank FRU → forward-filled from 45E1427.
    ns.append([None, None, None, None, None, None, None, "9JW154-038", "NA00", None, None, None, None])

    # A known-skipped sheet — must be ignored, NOT reported as unexpected.
    key = wb.create_sheet("Key Table")
    key.append(["Code", "Meaning"])
    key.append(["A1", "tested good"])

    path = tmp_path / "mini_fru_matrix.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture()
def parsed(mini_workbook):
    result = parse_workbook(mini_workbook)
    return result.links, {s.sheet: s for s in result.stats}


def _find(links, fru_norm=None, kind=None, related_norm=None):
    return [
        link
        for link in links
        if (fru_norm is None or link.fru_norm == fru_norm)
        and (kind is None or link.rel_kind == kind)
        and (related_norm is None or link.related_norm == related_norm)
    ]


class TestHygieneHelpers:
    def test_clean_strips_nbsp_and_collapses_whitespace(self):
        assert _clean("\xa0 SH20L07587 ") == "SH20L07587"
        assert _clean("a   b\xa0 c") == "a b c"

    @pytest.mark.parametrize("sentinel", ["N/ A", "N/A", "#N/A", "PENDIENTE", "None", "", "  ", "na", "Yes"])
    def test_clean_maps_sentinels_to_none(self, sentinel):
        assert _clean(sentinel) is None

    def test_pn_rejects_prose(self):
        assert _pn("Approved sub is 81Y9727 w/ same 11S PN") is None
        assert _pn("68Y7789") == "68Y7789"

    def test_pn_on_reject_called_for_prose_not_sentinels(self):
        rejected = []
        assert _pn("Approved sub is 81Y9727", on_reject=rejected.append) is None
        assert rejected == ["Approved sub is 81Y9727"]
        assert _pn("N/A", on_reject=rejected.append) is None
        assert _pn(None, on_reject=rejected.append) is None
        assert len(rejected) == 1  # sentinels/blanks are "no data", not parse failures

    def test_split_pns_comma_multivalue(self):
        assert _split_pns("39J3525, 53P5456") == ["39J3525", "53P5456"]

    def test_parse_carrier_slash_and_parenthetical(self):
        pns, note = _parse_carrier("SM10G01157 / 00FC544 (Blue hot-swap)")
        assert pns == ["SM10G01157", "00FC544"]
        assert note == "Blue hot-swap"

    def test_lenovo_fru_depads_and_strips_suffix(self):
        assert _lenovo_fru("0000000NV340_E00") == ("0000000NV340_E00", "00nv340")
        # 10-char PNs without zero padding are untouched.
        assert _lenovo_fru("SB17B49754") == ("SB17B49754", "sb17b49754")
        assert _lenovo_fru("  ") is None


class TestSheetCoverage:
    def test_full_workbook_covers_all_mapped_sheets(self, mini_workbook):
        result = parse_workbook(mini_workbook)
        assert len(result.stats) == len(PARSERS)
        assert result.missing_sheets == []
        assert result.unexpected_sheets == []  # "Key Table" is known-skipped

    def test_missing_mapped_sheet_is_fatal_by_default(self, mini_workbook, tmp_path):
        wb = openpyxl.load_workbook(mini_workbook)
        wb.remove(wb["Qlot as of 6.2025"])
        path = tmp_path / "missing.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="Qlot as of 6.2025"):
            parse_workbook(str(path))

    def test_allow_missing_sheets_escape_hatch(self, mini_workbook, tmp_path):
        wb = openpyxl.load_workbook(mini_workbook)
        wb.remove(wb["Qlot as of 6.2025"])
        path = tmp_path / "missing.xlsx"
        wb.save(path)
        result = parse_workbook(str(path), allow_missing_sheets=True)
        assert result.missing_sheets == ["Qlot as of 6.2025"]
        assert len(result.stats) == len(PARSERS) - 1

    def test_unexpected_sheet_reported_and_blocks_apply(self, mini_workbook, tmp_path):
        wb = openpyxl.load_workbook(mini_workbook)
        wb.create_sheet("Qlot as of 7.2025")  # renamed snapshot in a future revision
        path = tmp_path / "unexpected.xlsx"
        wb.save(path)
        result = parse_workbook(str(path))
        assert result.unexpected_sheets == ["Qlot as of 7.2025"]
        with pytest.raises(SystemExit, match="Qlot as of 7.2025"):
            main(str(path), apply=True)


class TestMainSheet:
    def test_kind_mapping(self, parsed):
        links, _ = parsed
        assert _find(links, "00aj001", FruLinkKind.IBM_11S, "68y7789")
        assert _find(links, "00aj001", FruLinkKind.OPTION, "00aj008")
        model = _find(links, "00aj001", FruLinkKind.MFG_MODEL)
        assert model and model[0].manufacturer == "Intel"
        assert model[0].machine == "x3650"
        assert model[0].series == "xSeries"
        assert "FC A2U4" in model[0].note and "ThinkServer" in model[0].note

    def test_sourcing_numbers_split_and_prose_dropped(self, parsed):
        links, _ = parsed
        sourcing = _find(links, kind=FruLinkKind.SOURCING_PN)
        assert {link.related_raw for link in sourcing if link.source_sheet == "Main"} == {"40K1107", "42C0432"}

    def test_carrier_split_with_parenthetical_note(self, parsed):
        links, _ = parsed
        trays = _find(links, "00aj002", FruLinkKind.TRAY)
        assert {t.related_raw for t in trays} == {"SM10G01157", "00FC544"}
        assert all("Blue hot-swap" in t.note for t in trays)
        single = _find(links, "00aj001", FruLinkKind.TRAY)
        assert single[0].related_raw == "41Y0708"
        assert "w/interposer 41Y0709" in single[0].note

    def test_nbsp_and_sentinels(self, parsed):
        links, _ = parsed
        # "\xa0 68Y7790 " cleaned; PENDIENTE manufacturer nulled; N/A option dropped.
        assert _find(links, "00aj002", FruLinkKind.IBM_11S, "68y7790")
        assert _find(links, "00aj002", FruLinkKind.MFG_MODEL)[0].manufacturer is None
        assert not _find(links, "00aj002", FruLinkKind.OPTION)
        assert not _find(links, "00aj002", FruLinkKind.TRAY_ALT)  # "N/ A"

    def test_self_edge_dropped(self, parsed):
        links, stats = parsed
        # Part Number == FRU → no self-referential ibm_11s link, counted as unparsed.
        assert not _find(links, "00aj003")
        s = stats["Main"]
        assert s.duplicate_links == 0
        assert s.unparsed_cells[FruLinkKind.IBM_11S.value] == 1

    def test_row_accounting(self, parsed):
        _, stats = parsed
        s = stats["Main"]
        assert s.rows == 5
        assert s.empty_rows == 1  # blank padding row
        assert s.skipped_rows == 1  # prose FRU
        # The prose sourcing cell splits into two non-PN tokens — both counted.
        assert s.unparsed_cells[FruLinkKind.SOURCING_PN.value] == 2


class TestQlotSheet:
    def test_drive_pn_only_with_qual_and_type_fallback(self, parsed):
        links, _ = parsed
        drive = _find(links, "00y2459", FruLinkKind.DRIVE_PN, "00y2430")
        assert drive
        d = drive[0]
        assert d.qual_status == "qlot approved"
        assert d.qual_date is None  # Qlot has no date column
        assert d.machine == "V3700"  # machine falls back to Type (no Brand column)
        assert d.note == "drop-in"
        # Qlot has no Model/Tray columns → drive_pn is the only kind emitted.
        assert not _find(links, "00y2459", FruLinkKind.MFG_MODEL)
        assert not _find(links, "00y2459", FruLinkKind.TRAY)


class TestGaborSheet:
    def test_drive_pn_carries_qual_context(self, parsed):
        links, _ = parsed
        drive = _find(links, "03gu649", FruLinkKind.DRIVE_PN, "00vn562")
        assert drive
        d = drive[0]
        assert d.qual_status == "qlot approved"
        assert str(d.qual_date) == "2024-03-14"
        assert d.manufacturer == "WDC"
        assert d.description == "18TB 3.5 HDD 7.2K SAS"
        assert d.machine == "V5000 Gen2"

    def test_tray_and_model_links(self, parsed):
        links, _ = parsed
        assert _find(links, "03gu649", FruLinkKind.TRAY, "01lj138")
        assert _find(links, "03gu649", FruLinkKind.MFG_MODEL, "wuh721818al4200")
        # Drive Model "#N/A" → no mfg_model link for the second FRU.
        assert not _find(links, "01ac602", FruLinkKind.MFG_MODEL)


class TestCzSheet:
    def test_oem_and_desc_column_positions(self, parsed):
        links, _ = parsed
        drive = _find(links, "01ej739", FruLinkKind.DRIVE_PN, "01ej710")[0]
        assert drive.manufacturer == "Seagate"  # OEM is column 6 (before Bare desc)
        assert drive.description == "8TB 3.5 HDD"  # FRU desc, column 8
        assert drive.machine == "V5000"  # machine falls back to Type (no Brand column)
        assert drive.qual_status == "qlot approved - Only EMEA"
        assert str(drive.qual_date) == "2024-05-01"
        model = _find(links, "01ej739", FruLinkKind.MFG_MODEL, "st8000nm0185")[0]
        assert model.description == "Bare 8TB NL-SAS"  # Bare desc is column 7
        assert _find(links, "01ej739", FruLinkKind.TRAY, "01ej900")


class TestCdcSheet:
    def test_fru_from_column_1_and_cdc_pending(self, parsed):
        links, _ = parsed
        drive = _find(links, "01gv870", FruLinkKind.DRIVE_PN, "01gv840")
        assert drive  # FRU comes from column 1 (column 0 is the FRU/DRIVE concat)
        d = drive[0]
        assert d.qual_status == CDC_PENDING
        assert d.description == "10TB 7.2K NL SAS drive"
        assert d.machine == "FlashSystem"
        assert d.note == "CDC platform: CDC4"
        # The concat column 0 must NOT be treated as the FRU.
        assert not _find(links, "01gv87001gv840")


class TestLenovoHddSheet:
    def test_lenovo_pn_and_fw_note(self, parsed):
        links, _ = parsed
        # Sole producer of the lenovo_pn kind.
        assert _find(links, "00yk010", FruLinkKind.LENOVO_PN, "sss7a12345")
        assert _find(links, "00yk010", FruLinkKind.IBM_11S, "00yk011")
        model = _find(links, "00yk010", FruLinkKind.MFG_MODEL, "st600mm0088")[0]
        assert model.manufacturer == "Seagate"
        assert model.note == "FW LK69"  # FW annotates the mfg_model link, not the tray
        assert model.machine == "SR650"
        assert model.series == "ThinkSystem"
        tray = _find(links, "00yk010", FruLinkKind.TRAY, "00e7600")[0]
        assert tray.note == "Gen3 tray"


class TestLenovoFruPnSheet:
    def test_fru_and_ppn_depadded(self, parsed):
        links, stats = parsed
        link = _find(links, "00nv340", FruLinkKind.LENOVO_PPN)[0]
        assert link.fru_raw == "0000000NV340_E00"  # raw keeps the padded original
        assert link.related_raw == "ESG0017964"
        assert link.note == "FRU-PPN"
        padded_ppn = _find(links, "sb17b49754", FruLinkKind.LENOVO_PPN)[0]
        assert padded_ppn.related_raw == "0000049Y4746"
        assert padded_ppn.related_norm == "49y4746"  # de-padded for joining

    def test_prose_fru_and_ppn_rejected(self, parsed):
        links, stats = parsed
        s = stats["Lenovo FRU-PN"]
        assert s.skipped_rows == 2  # blank FRU + prose FRU ("see SBB above")
        # The prose-FRU row's PPN must not leak in as a link.
        assert not _find(links, related_norm="esg0017965")
        # Valid FRU with a prose PPN: no link, counted as an unparsed cell.
        assert not _find(links, "00nv341")
        assert s.unparsed_cells[FruLinkKind.LENOVO_PPN.value] == 1


class TestLvnVpdSheet:
    def test_platform_brand_is_note_not_manufacturer(self, parsed):
        links, _ = parsed
        model = _find(links, "02jg542", FruLinkKind.MFG_MODEL, "mzilt3t8hbls")[0]
        assert model.manufacturer is None  # "ThinkSystem" is a platform label
        assert model.note == "Brand: ThinkSystem"
        options = _find(links, "02jg542", FruLinkKind.OPTION)
        assert {o.related_raw for o in options} == {"4XB7A14099", "4XB7A14100"}  # both option columns
        mtl = _find(links, "02jg542", FruLinkKind.SOURCING_PN, "sss7a23456")[0]
        assert "make-to-label" in mtl.note and "Brand: ThinkSystem" in mtl.note
        assert _find(links, "02jg542", FruLinkKind.TRAY, "01pg596")

    def test_non_platform_brand_is_manufacturer(self, parsed):
        links, _ = parsed
        model = _find(links, "02jg543", FruLinkKind.MFG_MODEL, "mzilt7t6hala")[0]
        assert model.manufacturer == "Samsung"
        assert model.note is None


class TestSeriesSheet:
    def test_hardware_links_and_compact_note(self, parsed):
        links, _ = parsed
        boards = _find(links, "00p1517", FruLinkKind.BOARD)
        assert {b.related_raw for b in boards} == {"39J3525", "53P5456"}  # comma split
        brackets = _find(links, "00p1517", FruLinkKind.BRACKET)
        assert {b.related_raw for b in brackets} == {"U3X", "53P5457"}
        assert _find(links, "00p1517", FruLinkKind.SCREWS, "05j7985")
        model = _find(links, "00p1517", FruLinkKind.MFG_MODEL)[0]
        assert model.manufacturer == "HITACHI/IBM"
        assert model.note == "9GB 3.5in 10K 80pin"
        assert model.series == "pSeries"

    def test_pendiente_model_skipped(self, parsed):
        links, _ = parsed
        assert not _find(links, "00p1518", FruLinkKind.MFG_MODEL)
        assert _find(links, "00p1518", FruLinkKind.IBM_11S, "71p7530")


class TestNSeriesSheet:
    def test_vendor_models_with_fw_notes(self, parsed):
        links, _ = parsed
        hitachi = _find(links, "45e1427", FruLinkKind.MFG_MODEL, "0a35002")[0]
        assert hitachi.manufacturer == "Hitachi"
        assert hitachi.note == "FW A90A"
        wd = _find(links, "45e1427", FruLinkKind.MFG_MODEL, "wd1002fbys05a6b0")[0]
        assert wd.manufacturer == "Western Digital"

    def test_forward_fill_continuation_row(self, parsed):
        links, _ = parsed
        seagate = _find(links, "45e1427", FruLinkKind.MFG_MODEL, "9jw154038")
        assert seagate and seagate[0].manufacturer == "Seagate"
        assert seagate[0].note == "FW NA00"

    def test_shuttle_dongle_screws(self, parsed):
        links, _ = parsed
        shuttles = _find(links, "45e1427", FruLinkKind.SHUTTLE)
        assert {s.related_raw for s in shuttles} == {"X298A-R5", "SP-298A-R5"}
        assert all("Shuttle type: DS14MK2ATA" in s.note for s in shuttles)
        assert _find(links, "45e1427", FruLinkKind.DONGLE, "6569502")
        assert _find(links, "45e1427", FruLinkKind.SCREWS, "83315xx")


class TestContextLengthCaps:
    def test_overlong_context_truncated_to_column_limits(self):
        # SQLite ignores VARCHAR lengths but PostgreSQL raises
        # StringDataRightTruncation — the emitter must cap at parse time.
        em = _Emitter("Main")
        em.add(
            "00AJ001",
            "00aj001",
            "ST9300603SS",
            FruLinkKind.MFG_MODEL,
            manufacturer="X" * 200,
            machine="M" * 200,
            series="S" * 100,
            qual_status="Q" * 100,
        )
        link = em.links[0]
        assert len(link.manufacturer) == 128 and link.manufacturer.endswith("…")
        assert len(link.machine) == 128 and link.machine.endswith("…")
        assert len(link.series) == 64 and link.series.endswith("…")
        assert len(link.qual_status) == 64 and link.qual_status.endswith("…")
        assert em.stats.unparsed_cells["manufacturer:truncated"] == 1
        assert em.stats.unparsed_cells["series:truncated"] == 1

    def test_within_limit_context_untouched(self):
        em = _Emitter("Main")
        em.add("00AJ001", "00aj001", "ST9300603SS", FruLinkKind.MFG_MODEL, manufacturer="Seagate")
        assert em.links[0].manufacturer == "Seagate"
        assert not em.stats.unparsed_cells


class TestUpsert:
    def _link(self, **overrides):
        base = dict(
            fru_raw="00AJ001",
            fru_norm="00aj001",
            related_raw="68Y7789",
            related_norm="68y7789",
            rel_kind=FruLinkKind.IBM_11S.value,
            source_sheet="Main",
        )
        base.update(overrides)
        return ParsedLink(**base)

    def test_insert_then_idempotent_reapply(self, db_session, mini_workbook):
        links = parse_workbook(mini_workbook).links
        inserted, updated = upsert_links(db_session, links, chunk_size=7)
        assert inserted == len(links)
        assert updated == 0
        assert db_session.query(FruLink).count() == len(links)

        inserted2, updated2 = upsert_links(db_session, links, chunk_size=7)
        assert inserted2 == 0
        assert updated2 == 0
        assert db_session.query(FruLink).count() == len(links)

    def test_update_refreshes_attributes(self, db_session):
        upsert_links(db_session, [self._link()])
        row = db_session.query(FruLink).one()
        assert row.manufacturer is None

        inserted, updated = upsert_links(db_session, [self._link(manufacturer="Intel", note="FC A2U4")])
        assert (inserted, updated) == (0, 1)
        row = db_session.query(FruLink).one()
        assert row.manufacturer == "Intel"
        assert row.note == "FC A2U4"

    def test_reapply_with_none_preserves_existing_attributes(self, db_session):
        # Additive-only contract: an attribute cleared in a newer workbook (None)
        # never nulls the previously ingested value.
        upsert_links(db_session, [self._link(manufacturer="Intel")])
        inserted, updated = upsert_links(db_session, [self._link(manufacturer=None)])
        assert (inserted, updated) == (0, 0)
        assert db_session.query(FruLink).one().manufacturer == "Intel"

    def test_absent_edges_survive_reingest(self, db_session):
        # Additive-only contract: edges missing from a newer workbook are NOT
        # deleted — withdrawn quals are a known limitation, not an accident.
        upsert_links(db_session, [self._link(), self._link(related_raw="99Y9999", related_norm="99y9999")])
        upsert_links(db_session, [self._link()])
        assert db_session.query(FruLink).count() == 2

    def test_unknown_rel_kind_rejected(self, db_session):
        # The bulk Core insert bypasses the ORM @validates hook, so upsert_links
        # re-validates rel_kind itself.
        with pytest.raises(ValueError):
            upsert_links(db_session, [self._link(rel_kind="not_a_kind")])
        assert db_session.query(FruLink).count() == 0
