# tests/test_quote_builder.py
"""tests/test_quote_builder.py — Quote Builder service, schemas, and endpoint tests.

Called by: pytest
Depends on: app.schemas.quote_builder, app.services.quote_builder_service, conftest.py
"""

from app.schemas.quote_builder import QuoteBuilderLine, QuoteBuilderSaveRequest


def test_builder_line_schema_valid():
    line = QuoteBuilderLine(
        requirement_id=1,
        offer_id=42,
        mpn="LM358DR",
        manufacturer="TI",
        qty=500,
        cost_price=0.24,
        sell_price=0.31,
        margin_pct=22.6,
    )
    assert line.mpn == "LM358DR"
    assert line.cost_price == 0.24


def test_builder_line_schema_optional_fields():
    line = QuoteBuilderLine(
        requirement_id=1,
        mpn="LM358DR",
        manufacturer="TI",
        qty=500,
        cost_price=0.24,
        sell_price=0.31,
        margin_pct=22.6,
    )
    assert line.offer_id is None
    assert line.lead_time is None
    assert line.notes is None


def test_builder_save_request_valid():
    req = QuoteBuilderSaveRequest(
        lines=[
            QuoteBuilderLine(
                requirement_id=1,
                mpn="LM358DR",
                manufacturer="TI",
                qty=500,
                cost_price=0.24,
                sell_price=0.31,
                margin_pct=22.6,
            )
        ],
        payment_terms="Net 30",
        shipping_terms="FCA",
        validity_days=7,
    )
    assert len(req.lines) == 1
    assert req.payment_terms == "Net 30"


def test_builder_save_request_empty_lines_rejected():
    import pytest as _pt

    with _pt.raises(Exception):
        QuoteBuilderSaveRequest(
            lines=[],
            payment_terms="Net 30",
        )


from app.services.quote_builder_service import apply_smart_defaults


def _mock_requirement(req_id, mpn, offers_count, target_qty=100, target_price=1.0):
    """Build a mock requirement dict as returned by get_builder_data."""
    offers = []
    for i in range(offers_count):
        offers.append(
            {
                "id": req_id * 100 + i,
                "vendor_name": f"Vendor{i}",
                "unit_price": 0.50 + i * 0.10,
                "qty_available": 500,
                "lead_time": "2 weeks",
                "date_code": "2024+",
                "condition": "new",
                "packaging": None,
                "moq": 100,
                "confidence": 0.95 if i == 0 else None,
                "notes": None,
            }
        )
    return {
        "requirement_id": req_id,
        "mpn": mpn,
        "manufacturer": "TI",
        "target_qty": target_qty,
        "target_price": target_price,
        "customer_pn": None,
        "date_codes": None,
        "condition": None,
        "packaging": None,
        "firmware": None,
        "hardware_codes": None,
        "sale_notes": None,
        "need_by_date": None,
        "offers": offers,
        "offer_count": offers_count,
        "status": "unknown",
        "selected_offer_id": None,
        "sell_price": None,
        "sell_price_manual": False,
        "buyer_notes": "",
        "pricing_history": None,
    }


def test_smart_defaults_single_offer_auto_decided():
    lines = [_mock_requirement(1, "LM358DR", 1)]
    apply_smart_defaults(lines)
    assert lines[0]["status"] == "decided"
    assert lines[0]["selected_offer_id"] == 100


def test_smart_defaults_multiple_offers_needs_review():
    lines = [_mock_requirement(2, "LM317T", 3)]
    apply_smart_defaults(lines)
    assert lines[0]["status"] == "needs_review"
    assert lines[0]["selected_offer_id"] is None


def test_smart_defaults_no_offers():
    lines = [_mock_requirement(3, "NE555P", 0)]
    apply_smart_defaults(lines)
    assert lines[0]["status"] == "no_offers"
    assert lines[0]["selected_offer_id"] is None


def test_smart_defaults_auto_pick_sets_sell_price():
    lines = [_mock_requirement(4, "SN74HC00N", 1)]
    apply_smart_defaults(lines)
    assert lines[0]["sell_price"] == lines[0]["offers"][0]["unit_price"]
    assert lines[0]["sell_price_manual"] is False


from app.services.quote_builder_service import build_excel_export


def test_excel_export_produces_valid_xlsx():
    line_items = [
        {
            "mpn": "LM358DR",
            "manufacturer": "TI",
            "qty": 500,
            "sell_price": 0.31,
            "lead_time": "2 weeks",
            "date_code": "2024+",
            "condition": "new",
            "packaging": "Tape & Reel",
            "moq": 100,
            "vendor_name": "DigiKey",
        }
    ]
    xlsx_bytes = build_excel_export(
        line_items=line_items,
        quote_number="Q-2026-0042",
        customer_name="Acme Electronics",
    )
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100
    assert xlsx_bytes[:2] == b"PK"


def test_excel_export_has_correct_columns():
    from io import BytesIO

    import openpyxl

    line_items = [
        {
            "mpn": "LM358DR",
            "manufacturer": "TI",
            "qty": 500,
            "sell_price": 0.31,
            "lead_time": "2 weeks",
            "date_code": "2024+",
            "condition": "new",
            "packaging": "T&R",
            "moq": 100,
            "vendor_name": "DigiKey",
        }
    ]
    xlsx_bytes = build_excel_export(line_items, "Q-2026-0042", "Acme")
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "MPN" in headers
    assert "Manufacturer" in headers
    assert "Qty" in headers
    assert "Unit Price" in headers
    assert "Extended Price" in headers
    assert "Lead Time" in headers
    assert "Vendor" in headers
    assert ws.cell(row=2, column=1).value == "LM358DR"


from app.models import Company, CustomerSite, Quote, Requirement, Requisition


def test_save_quote_from_builder_creates_quote(db_session, test_user):
    """Uses conftest fixtures for DB session and user."""
    from app.schemas.quote_builder import QuoteBuilderLine, QuoteBuilderSaveRequest
    from app.services.quote_builder_service import save_quote_from_builder

    # Seed a requisition with customer site
    company = Company(name="Acme Corp")
    db_session.add(company)
    db_session.flush()
    site = CustomerSite(company_id=company.id, site_name="HQ")
    db_session.add(site)
    db_session.flush()
    req = Requisition(name="Test Req", customer_site_id=site.id, created_by=test_user.id, status="active")
    db_session.add(req)
    db_session.flush()
    r1 = Requirement(requisition_id=req.id, primary_mpn="LM358DR", manufacturer="TI", target_qty=500)
    db_session.add(r1)
    db_session.commit()

    payload = QuoteBuilderSaveRequest(
        lines=[
            QuoteBuilderLine(
                requirement_id=r1.id,
                mpn="LM358DR",
                manufacturer="TI",
                qty=500,
                cost_price=0.24,
                sell_price=0.31,
                margin_pct=22.6,
            )
        ],
        payment_terms="Net 30",
    )
    result = save_quote_from_builder(db_session, req_id=req.id, payload=payload, user=test_user)
    assert result["ok"] is True
    assert "quote_id" in result
    assert "quote_number" in result

    # Verify Quote record exists
    quote = db_session.get(Quote, result["quote_id"])
    assert quote is not None
    assert quote.payment_terms == "Net 30"
    assert len(quote.line_items) == 1


def test_save_quote_revision(db_session, test_user):
    """Uses conftest fixtures for DB session and user."""
    from app.schemas.quote_builder import QuoteBuilderLine, QuoteBuilderSaveRequest
    from app.services.quote_builder_service import save_quote_from_builder

    # Seed
    company = Company(name="Acme Corp")
    db_session.add(company)
    db_session.flush()
    site = CustomerSite(company_id=company.id, site_name="HQ")
    db_session.add(site)
    db_session.flush()
    req = Requisition(name="Test Req", customer_site_id=site.id, created_by=test_user.id, status="active")
    db_session.add(req)
    db_session.flush()
    r1 = Requirement(requisition_id=req.id, primary_mpn="LM358DR", manufacturer="TI", target_qty=500)
    db_session.add(r1)
    db_session.commit()

    # First save
    payload1 = QuoteBuilderSaveRequest(
        lines=[
            QuoteBuilderLine(
                requirement_id=r1.id,
                mpn="LM358DR",
                manufacturer="TI",
                qty=500,
                cost_price=0.24,
                sell_price=0.31,
                margin_pct=22.6,
            )
        ],
    )
    result1 = save_quote_from_builder(db_session, req_id=req.id, payload=payload1, user=test_user)
    quote_id_1 = result1["quote_id"]

    # Second save (revision) — same quote_id passed
    payload2 = QuoteBuilderSaveRequest(
        lines=[
            QuoteBuilderLine(
                requirement_id=r1.id,
                mpn="LM358DR",
                manufacturer="TI",
                qty=500,
                cost_price=0.24,
                sell_price=0.40,
                margin_pct=40.0,
            )
        ],
        quote_id=quote_id_1,
    )
    result2 = save_quote_from_builder(db_session, req_id=req.id, payload=payload2, user=test_user)
    assert result2["ok"] is True
    assert result2["quote_id"] != quote_id_1  # New quote for revision

    # Old quote should be "revised"
    old_quote = db_session.get(Quote, quote_id_1)
    assert old_quote.status == "revised"


def test_builder_modal_endpoint_404_bad_req(client):
    resp = client.get("/v2/partials/quote-builder/99999")
    assert resp.status_code == 404


def test_builder_modal_opens_successfully(client, db_session, test_user):
    """Verify the modal endpoint returns 200 with valid data."""
    company = Company(name="Test Co")
    db_session.add(company)
    db_session.flush()
    site = CustomerSite(company_id=company.id, site_name="HQ")
    db_session.add(site)
    db_session.flush()
    req = Requisition(name="Test Req", customer_site_id=site.id, created_by=test_user.id, status="active")
    db_session.add(req)
    db_session.commit()
    resp = client.get(f"/v2/partials/quote-builder/{req.id}")
    assert resp.status_code == 200
    assert "Quote Builder" in resp.text


def test_builder_save_endpoint_rejects_empty_lines(client):
    resp = client.post("/v2/partials/quote-builder/1/save", json={"lines": []})
    assert resp.status_code == 422


def test_builder_excel_export_404_bad_quote(client):
    resp = client.get("/v2/partials/quote-builder/1/export/excel?quote_id=99999")
    assert resp.status_code == 404


def test_builder_pdf_export_404_bad_quote(client):
    resp = client.get("/v2/partials/quote-builder/1/export/pdf?quote_id=99999")
    assert resp.status_code == 404
