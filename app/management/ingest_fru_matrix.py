"""Ingest the IBM/Lenovo "FRU_PN_TRAY matrix" workbook into the fru_links table.

Usage: python -m app.management.ingest_fru_matrix <xlsx> [--apply] [--allow-missing-sheets]

DEFAULT is a dry run: parses every mapped sheet and reports sheet coverage
("sheets parsed X/Y"), per-sheet parsed/skipped/unparsed-cell counts, per-kind link
counts, and 10 sample links — writes nothing. A mapped sheet missing from the
workbook is FATAL unless --allow-missing-sheets is passed; workbook sheets that are
neither mapped nor in KNOWN_SKIPPED_SHEETS are reported as unexpected and block
--apply (a renamed date-stamped sheet must be re-mapped in PARSERS, not silently
dropped). With --apply it upserts in chunks inside a single transaction (insert new
edges, refresh context attributes on existing ones, keyed on fru_norm +
related_norm + rel_kind + source_sheet). Upserts are additive-only: edges absent
from a newer workbook are never deleted, and attributes cleared in the source never
null existing values.

Sheet → relationship mapping (kinds are FruLinkKind values):
  Main:               Part Number→ibm_11s, Option→option, Opt-PN→option_pn,
                      Model→mfg_model, Assembly→assembly,
                      Additional Sourcing Numbers→sourcing_pn (split),
                      Carrier→tray, Alternate Carrier→tray_alt
  Qlot:               Drivepn→drive_pn (+qual status)
  Gabor/CZ:           Drivepn→drive_pn (+qual status/date), Tray→tray,
                      Drive Model→mfg_model
  CDC NOT yet AP:     Drivepn→drive_pn with qual_status=CDC_PENDING
  Lenovo-HDD:         Part Number→ibm_11s, Lenovo/Idea PN→lenovo_pn,
                      Model→mfg_model (FW → note), Carrier→tray
  Lenovo FRU-PN:      PPN→lenovo_ppn (FRU de-padded: trailing _<letter><digits>
                      revision suffix, e.g. _E00, + zero padding)
  LVN VPD Mapping:    Option P/N→option, MFG Model→mfg_model,
                      Make-to-Label→sourcing_pn, Tray P/N→tray
  Series:             partno→ibm_11s, idmodelo→mfg_model, bracket(+alt)→bracket,
                      board(+alt)→board, screws→screws
  NSeries(NetApp):    per-vendor MFG P/N→mfg_model, Shuttle→shuttle, Dongle→dongle,
                      Dongle Screw→screws (FRU forward-filled down blank rows)
Skipped sheets (KNOWN_SKIPPED_SHEETS): "11s Sub Check", "HD Matrix Template",
"Lenovo HD-SD Template" (lookup tools/templates) and "Key Table" (test/disposition
codes, no part links).

Called by: admin manually after receiving an updated FRU matrix workbook
Depends on: openpyxl, app.models.FruLink, app.constants.FruLinkKind/CDC_PENDING,
            app.utils.normalization.normalize_mpn_key
"""

import argparse
import re
from collections import Counter
from dataclasses import dataclass, field, fields, replace
from datetime import date, datetime
from functools import lru_cache
from typing import Callable, Iterable

import sqlalchemy as sa
from loguru import logger

from app.constants import CDC_PENDING, FruLinkKind
from app.utils.normalization import normalize_mpn_key

# ── Cell hygiene ──────────────────────────────────────────────────────

# Values that mean "no data" anywhere in the workbook (case-insensitive).
# "yes"/"no"/"na"/"x" appear as filler in PN columns (e.g. Lenovo-HDD Carrier="Yes").
_SENTINELS = {"", "n/a", "n/ a", "#n/a", "pendiente", "none", "na", "yes", "no", "x", "-", "?"}

# A plausible part number: starts alphanumeric, no whitespace, 3-40 chars.
# Prose cells ("Approved sub is 81Y9727 w/ same 11S PN ...") fail this and are dropped.
_PN_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9./+_-]{2,39}$")

_PAREN_RE = re.compile(r"\(([^)]*)\)")
_LENOVO_FRU_SUFFIX_RE = re.compile(r"_[A-Za-z]\d{1,3}$")  # e.g. "_E00"

# Lenovo platform/series labels that show up in Brand columns — NOT manufacturers.
_LENOVO_PLATFORM_BRANDS = {"lenovo", "lenovo storage", "system x", "thinksystem", "thinkserver", "ideapad"}


def _clean(value) -> str | None:
    """Trim, de-nbsp, collapse whitespace; map sentinel values to None."""
    if value is None:
        return None
    s = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    if s.lower() in _SENTINELS:
        return None
    return s


def _pn(value, on_reject: Callable[[str], None] | None = None) -> str | None:
    """A single plausible part number, or None (sentinels, prose, blanks).

    ``on_reject`` is called with the cleaned value when a NON-EMPTY cell fails the
    plausibility check (so callers can count dropped cells; sentinels don't count).
    """
    s = _clean(value)
    if not s:
        return None
    if not _PN_RE.match(s):
        if on_reject is not None:
            on_reject(s)
        return None
    return s


def _split_pns(value, seps: str = r"[,;\n]", on_reject: Callable[[str], None] | None = None) -> list[str]:
    """Split a multi-value cell and keep only plausible part numbers."""
    s = _clean(value)
    if not s:
        return []
    return [p for p in (_pn(tok, on_reject) for tok in re.split(seps, s)) if p]


def _parse_carrier(value, on_reject: Callable[[str], None] | None = None) -> tuple[list[str], str | None]:
    """Carrier cells: split on '/' (and , ; newline), parentheticals become the note.

    "SM10G01157 / 00FC544 (Blue hot-swap)" → (["SM10G01157", "00FC544"], "Blue hot-swap")
    """
    s = _clean(value)
    if not s:
        return [], None
    notes = [n.strip() for n in _PAREN_RE.findall(s) if n.strip()]
    bare = _PAREN_RE.sub(" ", s)
    pns = [p for p in (_pn(tok, on_reject) for tok in re.split(r"[/,;\n]", bare)) if p]
    return pns, ("; ".join(notes) or None)


def _lenovo_fru(value) -> tuple[str, str] | None:
    """Lenovo FRU-PN FRU cell → (raw, norm). "0000000NV340_E00" → norm of "00NV340".

    Strips the trailing ``_<letter><digits>`` revision suffix (e.g. ``_E00``), then
    the zero-padding group down to the 7-char FRU (only when everything before the
    last 7 chars is zeros — 10-char PNs like "SB17B49754" are left alone).
    """
    s = _clean(value)
    if not s:
        return None
    base = _LENOVO_FRU_SUFFIX_RE.sub("", s)
    if len(base) > 7 and not base[: len(base) - 7].strip("0"):
        base = base[-7:]
    norm = normalize_mpn_key(base)
    if len(norm) < 3:
        return None
    return s, norm


def _date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _num(value) -> str | None:
    """Compact number for note strings: 9.0 → "9", 3.5 → "3.5"."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = _clean(value)
    return s


def _join_notes(*parts: str | None) -> str | None:
    return "; ".join(p for p in parts if p) or None


def _pad(row: tuple, width: int) -> tuple:
    """read_only rows can be shorter than the header — pad with None."""
    return row + (None,) * (width - len(row)) if len(row) < width else row


def _is_empty(row: tuple) -> bool:
    """True when every cell is None/whitespace (trailing formatting rows)."""
    return not any(c is not None and str(c).strip() for c in row)


# ── Parsed link record ────────────────────────────────────────────────


@dataclass
class ParsedLink:
    """One FRU↔PN edge parsed from a sheet (mirrors FruLink columns)."""

    fru_raw: str
    fru_norm: str
    related_raw: str
    related_norm: str
    rel_kind: str
    source_sheet: str
    manufacturer: str | None = None
    description: str | None = None
    series: str | None = None
    machine: str | None = None
    qual_status: str | None = None
    qual_date: date | None = None
    note: str | None = None

    @property
    def key(self) -> tuple[str, str, str, str]:
        return (self.fru_norm, self.related_norm, self.rel_kind, self.source_sheet)


@dataclass
class SheetStats:
    """Dry-run/apply report numbers for one sheet."""

    sheet: str
    rows: int = 0  # rows below the header scanned (incl. blank)
    empty_rows: int = 0  # fully blank rows (sheet formatting padding)
    skipped_rows: int = 0  # rows with data but no usable FRU
    duplicate_links: int = 0  # parsed edges collapsed into an existing key
    kinds: Counter = field(default_factory=Counter)  # unique links per rel_kind
    # Non-empty cells/tokens dropped by validation (failed _PN_RE, normalized to
    # nothing, self-edge) or truncated to a column limit — keyed per kind/column.
    # Distinguishes "column was empty" from "column format changed" in the report.
    unparsed_cells: Counter = field(default_factory=Counter)


_IDENTITY_FIELDS = {"fru_raw", "fru_norm", "related_raw", "related_norm", "rel_kind", "source_sheet"}


@lru_cache(maxsize=1)
def _context_limits() -> dict[str, int]:
    """Length caps for bounded String context columns, derived from the FruLink model.

    description/note are Text (unbounded). Keeping these in sync with the model means a
    workbook cell can never blow a VARCHAR limit on PostgreSQL at --apply time (SQLite
    ignores lengths, so tests alone would never catch it).
    """
    from app.models.fru_link import FruLink

    return {
        name: FruLink.__table__.columns[name].type.length
        for name in ("manufacturer", "series", "machine", "qual_status")
    }


class _Emitter:
    """Per-sheet link collector: validates, dedups on the unique key, coalesces."""

    _CONTEXT_FIELDS = [f.name for f in fields(ParsedLink) if f.name not in _IDENTITY_FIELDS]
    _DEBUG_SAMPLES = 8  # max rejected raw values logged per sheet (audit trail)

    def __init__(self, sheet: str):
        self.sheet = sheet
        self.stats = SheetStats(sheet=sheet)
        self._by_key: dict[tuple, ParsedLink] = {}

    def scan(self, row: tuple) -> bool:
        """Count the row; returns False (and counts it as empty) for blank rows."""
        self.stats.rows += 1
        if _is_empty(row):
            self.stats.empty_rows += 1
            return False
        return True

    def _reject(self, column: str, raw: str, reason: str = "unparsable") -> None:
        """Count a dropped/altered non-empty value so the report can surface it."""
        self.stats.unparsed_cells[column] += 1
        if sum(self.stats.unparsed_cells.values()) <= self._DEBUG_SAMPLES:
            logger.debug("[{}] {} {} value: {!r}", self.sheet, reason, column, raw)

    def pn(self, value, column: str) -> str | None:
        """_pn() that counts non-empty cells failing the plausibility check."""
        return _pn(value, on_reject=lambda raw: self._reject(column, raw))

    def split_pns(self, value, column: str, seps: str = r"[,;\n]") -> list[str]:
        """_split_pns() that counts non-conforming tokens."""
        return _split_pns(value, seps, on_reject=lambda raw: self._reject(column, raw))

    def parse_carrier(self, value, column: str) -> tuple[list[str], str | None]:
        """_parse_carrier() that counts non-conforming tokens."""
        return _parse_carrier(value, on_reject=lambda raw: self._reject(column, raw))

    def add(
        self,
        fru_raw: str,
        fru_norm: str,
        related: str | None,
        kind: FruLinkKind,
        related_norm_override: str | None = None,
        **context,
    ) -> None:
        if not related:
            return
        related_norm = related_norm_override or normalize_mpn_key(related)
        if not related_norm or related_norm == fru_norm:
            reason = "self-edge" if related_norm else "empty-norm"
            self._reject(kind.value, related, reason=reason)
            return  # unusable or self-edge
        # Cap bounded context columns so --apply can never hit a Postgres
        # StringDataRightTruncation (truncate with an ellipsis, count it).
        for name, limit in _context_limits().items():
            val = context.get(name)
            if isinstance(val, str) and len(val) > limit:
                context[name] = val[: limit - 1] + "…"
                self._reject(f"{name}:truncated", val, reason="overlength")
        link = ParsedLink(
            fru_raw=fru_raw,
            fru_norm=fru_norm,
            related_raw=related,
            related_norm=related_norm,
            rel_kind=kind.value,
            source_sheet=self.sheet,
            **context,
        )
        existing = self._by_key.get(link.key)
        if existing is None:
            self._by_key[link.key] = link
            self.stats.kinds[kind.value] += 1
        else:
            # Coalesce: fill attributes the first occurrence was missing.
            fills = {
                name: getattr(link, name)
                for name in self._CONTEXT_FIELDS
                if getattr(existing, name) is None and getattr(link, name) is not None
            }
            if fills:
                self._by_key[link.key] = replace(existing, **fills)
            self.stats.duplicate_links += 1

    @property
    def links(self) -> list[ParsedLink]:
        return list(self._by_key.values())


# ── Per-sheet parsers ─────────────────────────────────────────────────


def parse_main(ws, sheet: str) -> _Emitter:
    """Main: FRU | Part Number | Option | Opt-PN | Model | Assembly | Addl Sourcing |
    Manufactured by | Carrier | Alternate Carrier | Machine | Feature Code | Comment | Series."""
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 14)
        if not em.scan(row):
            continue
        fru = _pn(row[0])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        fru_norm = normalize_mpn_key(fru)
        manufacturer = _clean(row[7])
        ctx = {
            "machine": _clean(row[10]),
            "series": _clean(row[13]),
            "note": _join_notes(f"FC {_clean(row[11])}" if _clean(row[11]) else None, _clean(row[12])),
        }
        em.add(fru, fru_norm, em.pn(row[1], FruLinkKind.IBM_11S.value), FruLinkKind.IBM_11S, **ctx)
        em.add(fru, fru_norm, em.pn(row[2], FruLinkKind.OPTION.value), FruLinkKind.OPTION, **ctx)
        em.add(fru, fru_norm, em.pn(row[3], FruLinkKind.OPTION_PN.value), FruLinkKind.OPTION_PN, **ctx)
        em.add(
            fru,
            fru_norm,
            em.pn(row[4], FruLinkKind.MFG_MODEL.value),
            FruLinkKind.MFG_MODEL,
            manufacturer=manufacturer,
            **ctx,
        )
        em.add(fru, fru_norm, em.pn(row[5], FruLinkKind.ASSEMBLY.value), FruLinkKind.ASSEMBLY, **ctx)
        for pn in em.split_pns(row[6], FruLinkKind.SOURCING_PN.value):
            em.add(fru, fru_norm, pn, FruLinkKind.SOURCING_PN, **ctx)
        for col, kind in ((8, FruLinkKind.TRAY), (9, FruLinkKind.TRAY_ALT)):
            pns, carrier_note = em.parse_carrier(row[col], kind.value)
            for pn in pns:
                em.add(fru, fru_norm, pn, kind, **{**ctx, "note": _join_notes(ctx["note"], carrier_note)})
    return em


def _parse_qual_sheet(ws, sheet: str, cols: dict[str, int | None], qual_override: str | None = None) -> _Emitter:
    """Shared parser for the qual-list sheets (Qlot / Gabor / CZ / CDC).

    ``cols`` maps logical fields → column index (None when the sheet lacks it).
    """
    em = _Emitter(sheet)
    width = max(c for c in cols.values() if c is not None) + 1

    def col(row, name):
        idx = cols.get(name)
        return row[idx] if idx is not None else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, width)
        if not em.scan(row):
            continue
        fru = _pn(col(row, "fru"))
        if not fru:
            em.stats.skipped_rows += 1
            continue
        fru_norm = normalize_mpn_key(fru)
        oem = _clean(col(row, "oem"))
        machine = _clean(col(row, "brand")) or _clean(col(row, "type"))
        qual = qual_override or _clean(col(row, "qual"))
        qual_date = _date(col(row, "qual_date"))
        fru_desc = _clean(col(row, "fru_desc"))
        bare_desc = _clean(col(row, "bare_desc"))
        note = _join_notes(bare_desc, _clean(col(row, "comment")))
        em.add(
            fru,
            fru_norm,
            em.pn(col(row, "drive_pn"), FruLinkKind.DRIVE_PN.value),
            FruLinkKind.DRIVE_PN,
            manufacturer=oem,
            description=fru_desc,
            machine=machine,
            qual_status=qual,
            qual_date=qual_date,
            note=note,
        )
        em.add(
            fru,
            fru_norm,
            em.pn(col(row, "model"), FruLinkKind.MFG_MODEL.value),
            FruLinkKind.MFG_MODEL,
            manufacturer=oem,
            description=bare_desc,
            machine=machine,
            qual_status=qual,
            qual_date=qual_date,
        )
        em.add(fru, fru_norm, em.pn(col(row, "tray"), FruLinkKind.TRAY.value), FruLinkKind.TRAY, machine=machine)
    return em


def parse_qlot(ws, sheet: str) -> _Emitter:
    """Qlot as of 6.2025: FRU/DRIVE | FRU | Drivepn | Type | IW CSP qual | comment."""
    return _parse_qual_sheet(
        ws, sheet, {"fru": 1, "drive_pn": 2, "type": 3, "qual": 4, "comment": 5, "model": None, "tray": None}
    )


def parse_gabor(ws, sheet: str) -> _Emitter:
    """Gabor 11.13.25: ...

    | qual | date | Brand | Bare desc | OEM | FRU desc | Tray | Model | _ | CDC.
    """
    return _parse_qual_sheet(
        ws,
        sheet,
        {
            "fru": 1,
            "drive_pn": 2,
            "type": 3,
            "qual": 4,
            "qual_date": 5,
            "brand": 6,
            "bare_desc": 7,
            "oem": 8,
            "fru_desc": 9,
            "tray": 10,
            "model": 11,
        },
    )


def parse_cz(ws, sheet: str) -> _Emitter:
    """CZ ONLY Testing: FRU/DRIVE | FRU | Drivepn | Type | qual | date | OEM |
    Bare desc | FRU desc | Tray | Model.

    Like Gabor but with no Brand column and OEM moved before Bare desc (qual is
    'qlot approved - Only EMEA'); machine falls back to Type.
    """
    return _parse_qual_sheet(
        ws,
        sheet,
        {
            "fru": 1,
            "drive_pn": 2,
            "type": 3,
            "qual": 4,
            "qual_date": 5,
            "oem": 6,
            "bare_desc": 7,
            "fru_desc": 8,
            "tray": 9,
            "model": 10,
        },
    )


def parse_cdc(ws, sheet: str) -> _Emitter:
    """CDC NOT yet AP Article Required: FRU/DRIVE | FRU | Drivepn | model(prose) |
    BRAND | CDC Platform (leading FRU/DRIVE concat column like the other qual sheets).

    Rows are pending-qualification → qual_status CDC_PENDING; the prose model text
    becomes the description.
    """
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 6)
        if not em.scan(row):
            continue
        fru = _pn(row[1])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        platform = _clean(row[5])
        em.add(
            fru,
            normalize_mpn_key(fru),
            em.pn(row[2], FruLinkKind.DRIVE_PN.value),
            FruLinkKind.DRIVE_PN,
            description=_clean(row[3]),
            machine=_clean(row[4]),
            qual_status=CDC_PENDING,
            note=f"CDC platform: {platform}" if platform else None,
        )
    return em


def parse_lenovo_hdd(ws, sheet: str) -> _Emitter:
    """Lenovo-HDD: FRU | Part Number | Lenovo/Idea PN | Opt-PN | Model | ... | Manufactured by |
    Description | Carrier | Machine | FW | ... | Series."""
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 15)
        if not em.scan(row):
            continue
        fru = _pn(row[0])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        fru_norm = normalize_mpn_key(fru)
        ctx = {"description": _clean(row[8]), "machine": _clean(row[10]), "series": _clean(row[14])}
        fw = _clean(row[11])
        em.add(fru, fru_norm, em.pn(row[1], FruLinkKind.IBM_11S.value), FruLinkKind.IBM_11S, **ctx)
        em.add(fru, fru_norm, em.pn(row[2], FruLinkKind.LENOVO_PN.value), FruLinkKind.LENOVO_PN, **ctx)
        em.add(
            fru,
            fru_norm,
            em.pn(row[4], FruLinkKind.MFG_MODEL.value),
            FruLinkKind.MFG_MODEL,
            manufacturer=_clean(row[7]),
            note=f"FW {fw}" if fw else None,
            **ctx,
        )
        pns, carrier_note = em.parse_carrier(row[9], FruLinkKind.TRAY.value)
        for pn in pns:
            em.add(fru, fru_norm, pn, FruLinkKind.TRAY, note=carrier_note, **ctx)
    return em


def parse_lenovo_fru_pn(ws, sheet: str) -> _Emitter:
    """Lenovo FRU-PN: BOM Type | SBB Change Type | SBB | FRU | Qty | PPN | Last Modified.

    PPN values use the same SAP zero-padding as the FRU column ("0000000WG788" is
    "00WG788"), so the related side is de-padded for its norm too — raw keeps the
    original padded value. Both sides are gated through _pn() first (the padded
    "0000000NV340_E00" form passes _PN_RE), so prose cells are rejected exactly
    like every other sheet's PN columns.
    """
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 6)
        if not em.scan(row):
            continue
        fru_pair = _lenovo_fru(row[3]) if _pn(row[3]) else None
        if not fru_pair:
            em.stats.skipped_rows += 1
            continue
        fru_raw, fru_norm = fru_pair
        ppn_cell = em.pn(row[5], FruLinkKind.LENOVO_PPN.value)
        ppn_pair = _lenovo_fru(ppn_cell) if ppn_cell else None
        if ppn_pair:
            ppn_raw, ppn_norm = ppn_pair
            em.add(
                fru_raw, fru_norm, ppn_raw, FruLinkKind.LENOVO_PPN, note=_clean(row[0]), related_norm_override=ppn_norm
            )
    return em


def parse_lvn_vpd(ws, sheet: str) -> _Emitter:
    """LVN VPD Mapping: Brand | ... | Option P/N | FRU P/N | Option P/N | MFG Model |
    Make-to-Label P/N | Tray P/N | Description | ...

    Brand is a Lenovo platform label ("System x", "ThinkSystem") → note, not manufacturer;
    only a non-platform Brand would be treated as a manufacturer.
    """
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 11)
        if not em.scan(row):
            continue
        fru = _pn(row[5])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        fru_norm = normalize_mpn_key(fru)
        brand = _clean(row[0])
        is_platform = brand is not None and brand.lower() in _LENOVO_PLATFORM_BRANDS
        manufacturer = brand if brand and not is_platform else None
        ctx = {
            "description": _clean(row[10]),
            "note": f"Brand: {brand}" if is_platform else None,
        }
        em.add(fru, fru_norm, em.pn(row[4], FruLinkKind.OPTION.value), FruLinkKind.OPTION, **ctx)
        em.add(fru, fru_norm, em.pn(row[6], FruLinkKind.OPTION.value), FruLinkKind.OPTION, **ctx)
        em.add(
            fru,
            fru_norm,
            em.pn(row[7], FruLinkKind.MFG_MODEL.value),
            FruLinkKind.MFG_MODEL,
            manufacturer=manufacturer,
            **ctx,
        )
        em.add(
            fru,
            fru_norm,
            em.pn(row[8], FruLinkKind.SOURCING_PN.value),
            FruLinkKind.SOURCING_PN,
            description=ctx["description"],
            note=_join_notes("make-to-label", ctx["note"]),
        )
        em.add(fru, fru_norm, em.pn(row[9], FruLinkKind.TRAY.value), FruLinkKind.TRAY, **ctx)
    return em


def parse_series(ws, sheet: str) -> _Emitter:
    """Series: fru | partno | idmodelo | Manufacturer | Capacity | Size | rpm | Pins |
    partnosg | bracket | bracket_alt | board | board_alt | screws | ... | Brand(series)."""
    em = _Emitter(sheet)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = _pad(row, 18)
        if not em.scan(row):
            continue
        fru = _pn(row[0])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        fru_norm = normalize_mpn_key(fru)
        size_bits = [
            f"{_num(row[4])}GB" if _num(row[4]) else None,
            f"{_num(row[5])}in" if _num(row[5]) else None,
            f"{_num(row[6])}K" if _num(row[6]) else None,
            f"{_num(row[7])}pin" if _num(row[7]) else None,
        ]
        ctx = {"series": _clean(row[17]), "note": " ".join(b for b in size_bits if b) or None}
        em.add(fru, fru_norm, em.pn(row[1], FruLinkKind.IBM_11S.value), FruLinkKind.IBM_11S, **ctx)
        em.add(
            fru,
            fru_norm,
            em.pn(row[2], FruLinkKind.MFG_MODEL.value),
            FruLinkKind.MFG_MODEL,
            manufacturer=_clean(row[3]),
            **ctx,
        )
        for col, kind in (
            (9, FruLinkKind.BRACKET),
            (10, FruLinkKind.BRACKET),
            (11, FruLinkKind.BOARD),
            (12, FruLinkKind.BOARD),
            (13, FruLinkKind.SCREWS),
        ):
            for pn in em.split_pns(row[col], kind.value):
                em.add(fru, fru_norm, pn, kind, **ctx)
    return em


def parse_nseries(ws, sheet: str) -> _Emitter:
    """NSeries(NetApp): FRU | Description | 3X5 | Shuttle Type | Shuttle/Screw P/N |
    Dongle P/N | Dongle Screw P/N | Seagate MFG+FW | Hitachi MFG+FW | WD MFG+FW.

    Two header rows; a blank FRU cell continues the FRU above (forward-fill).
    """
    em = _Emitter(sheet)
    fru: str | None = None
    fru_norm = ""
    description: str | None = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        row = _pad(row, 13)
        if not em.scan(row):
            continue
        row_fru = _pn(row[0])
        if row_fru:
            fru, fru_norm, description = row_fru, normalize_mpn_key(row_fru), _clean(row[1])
        if not fru:
            em.stats.skipped_rows += 1
            continue
        shuttle_type = _clean(row[3])
        for pn in em.split_pns(row[4], FruLinkKind.SHUTTLE.value, seps=r"[/,;\n]"):
            em.add(
                fru,
                fru_norm,
                pn,
                FruLinkKind.SHUTTLE,
                description=description,
                note=f"Shuttle type: {shuttle_type}" if shuttle_type else None,
            )
        for pn in em.split_pns(row[5], FruLinkKind.DONGLE.value, seps=r"[/,;\n]"):
            em.add(fru, fru_norm, pn, FruLinkKind.DONGLE, description=description)
        for pn in em.split_pns(row[6], FruLinkKind.SCREWS.value, seps=r"[/,;\n]"):
            em.add(fru, fru_norm, pn, FruLinkKind.SCREWS, description=description)
        for pn_col, fw_col, mfr in ((7, 8, "Seagate"), (9, 10, "Hitachi"), (11, 12, "Western Digital")):
            fw = _clean(row[fw_col])
            for pn in em.split_pns(row[pn_col], FruLinkKind.MFG_MODEL.value, seps=r"[/,;\n]"):
                em.add(
                    fru,
                    fru_norm,
                    pn,
                    FruLinkKind.MFG_MODEL,
                    manufacturer=mfr,
                    description=description,
                    note=f"FW {fw}" if fw else None,
                )
    return em


PARSERS = [
    ("Main", parse_main),
    ("Qlot as of 6.2025", parse_qlot),
    ("Gabor 11.13.25", parse_gabor),
    ("CZ ONLY Testing", parse_cz),
    ("CDC NOT yet AP Article Required", parse_cdc),
    ("Lenovo-HDD", parse_lenovo_hdd),
    ("Lenovo FRU-PN", parse_lenovo_fru_pn),
    ("LVN VPD Mapping", parse_lvn_vpd),
    ("Series", parse_series),
    ("NSeries(NetApp)", parse_nseries),
]

# Workbook sheets that are deliberately NOT ingested (lookup tools, templates,
# disposition-code tables). Anything else not in PARSERS is unexpected and blocks
# --apply — a renamed date-stamped sheet must be re-mapped, never silently dropped.
KNOWN_SKIPPED_SHEETS = {"11s Sub Check", "HD Matrix Template", "Lenovo HD-SD Template", "Key Table"}


@dataclass
class WorkbookParse:
    """parse_workbook() result: links + per-sheet stats + sheet-coverage accounting."""

    links: list[ParsedLink]
    stats: list[SheetStats]
    missing_sheets: list[str]  # mapped sheets absent from the workbook (allow_missing_sheets only)
    unexpected_sheets: list[str]  # workbook sheets neither mapped nor known-skipped


def parse_workbook(path: str, allow_missing_sheets: bool = False) -> WorkbookParse:
    """Parse every mapped sheet of the workbook into deduplicated links + stats.

    Raises ValueError when a mapped sheet is missing (the four date-stamped sheet
    names WILL change in future workbook revisions — silently skipping one would
    drop a whole relationship class), unless ``allow_missing_sheets`` is set for
    deliberately partial workbooks.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    links: list[ParsedLink] = []
    stats: list[SheetStats] = []
    try:
        sheetnames = set(wb.sheetnames)
        missing = [name for name, _ in PARSERS if name not in sheetnames]
        if missing and not allow_missing_sheets:
            raise ValueError(
                f"Mapped sheet(s) missing from workbook: {missing} — if a sheet was renamed, "
                "update PARSERS; for a deliberately partial workbook pass --allow-missing-sheets"
            )
        unexpected = sorted(sheetnames - {name for name, _ in PARSERS} - KNOWN_SKIPPED_SHEETS)
        for sheet_name, parser in PARSERS:
            if sheet_name not in sheetnames:
                logger.warning("Sheet {!r} not found in workbook — skipping (--allow-missing-sheets)", sheet_name)
                continue
            emitter = parser(wb[sheet_name], sheet_name)
            links.extend(emitter.links)
            stats.append(emitter.stats)
    finally:
        wb.close()
    return WorkbookParse(links=links, stats=stats, missing_sheets=missing, unexpected_sheets=unexpected)


# ── Apply (chunked upsert) ────────────────────────────────────────────

_UPDATABLE = [
    "fru_raw",
    "related_raw",
    "manufacturer",
    "description",
    "series",
    "machine",
    "qual_status",
    "qual_date",
    "note",
]


def upsert_links(db, links: Iterable[ParsedLink], chunk_size: int = 1000) -> tuple[int, int]:
    """Insert new edges / refresh attributes on existing ones, in ONE transaction.

    A single commit at the end makes --apply all-or-nothing: a mid-run failure
    (IntegrityError, connection drop) rolls everything back instead of leaving an
    unreported half-applied state. Additive-only by design: edges absent from the
    parsed links are never deleted, and attributes cleared in the source (None)
    never null existing values — withdrawn qual statuses survive re-ingest.

    The bulk insert is a Core statement that bypasses the FruLink @validates hook,
    so rel_kind is re-validated here against FruLinkKind (raises ValueError).

    Returns (inserted, updated).
    """
    from app.models import FruLink

    links = list(links)
    inserted = updated = 0
    chunk_count = -(-len(links) // chunk_size) if links else 0
    for chunk_idx, start in enumerate(range(0, len(links), chunk_size), start=1):
        chunk = links[start : start + chunk_size]
        fru_norms = {link.fru_norm for link in chunk}
        existing = {
            (r.fru_norm, r.related_norm, r.rel_kind, r.source_sheet): r
            for r in db.execute(sa.select(FruLink).where(FruLink.fru_norm.in_(fru_norms))).scalars().all()
        }
        to_insert = []
        for link in chunk:
            row = existing.get(link.key)
            if row is None:
                to_insert.append(
                    {
                        "fru_raw": link.fru_raw,
                        "fru_norm": link.fru_norm,
                        "related_raw": link.related_raw,
                        "related_norm": link.related_norm,
                        "rel_kind": FruLinkKind(link.rel_kind).value,  # Core insert skips @validates
                        "source_sheet": link.source_sheet,
                        "manufacturer": link.manufacturer,
                        "description": link.description,
                        "series": link.series,
                        "machine": link.machine,
                        "qual_status": link.qual_status,
                        "qual_date": link.qual_date,
                        "note": link.note,
                    }
                )
            else:
                changed = False
                for attr in _UPDATABLE:
                    new = getattr(link, attr)
                    if new is not None and getattr(row, attr) != new:
                        setattr(row, attr, new)
                        changed = True
                if changed:
                    updated += 1
        if to_insert:
            db.execute(sa.insert(FruLink), to_insert)
            inserted += len(to_insert)
        logger.debug("chunk {}/{} staged: {} inserted, {} updated so far", chunk_idx, chunk_count, inserted, updated)
    db.commit()
    return inserted, updated


# ── CLI ───────────────────────────────────────────────────────────────


def _report(parsed: WorkbookParse) -> None:
    links, stats = parsed.links, parsed.stats
    logger.info("Sheets parsed: {}/{} mapped", len(stats), len(PARSERS))
    if parsed.missing_sheets:
        logger.warning("Missing mapped sheets (allowed): {}", parsed.missing_sheets)
    if parsed.unexpected_sheets:
        logger.error(
            "Unexpected sheets (not mapped, not in KNOWN_SKIPPED_SHEETS — renamed sheet? --apply refused): {}",
            parsed.unexpected_sheets,
        )
    for s in stats:
        kind_summary = ", ".join(f"{k}={n}" for k, n in sorted(s.kinds.items())) or "none"
        logger.info(
            "[{}] rows={} empty={} skipped={} dup_links={} links={} ({})",
            s.sheet,
            s.rows,
            s.empty_rows,
            s.skipped_rows,
            s.duplicate_links,
            sum(s.kinds.values()),
            kind_summary,
        )
        if s.unparsed_cells:
            logger.info(
                "[{}] unparsed cells: {}", s.sheet, ", ".join(f"{k}={n}" for k, n in sorted(s.unparsed_cells.items()))
            )
    totals: Counter = Counter()
    for s in stats:
        totals.update(s.kinds)
    logger.info("TOTAL links={} by kind: {}", len(links), ", ".join(f"{k}={n}" for k, n in sorted(totals.items())))
    step = max(1, len(links) // 10)
    for link in links[::step][:10]:
        logger.info(
            "sample: [{}] {} -{}-> {} (mfr={}, qual={})",
            link.source_sheet,
            link.fru_raw,
            link.rel_kind,
            link.related_raw,
            link.manufacturer,
            link.qual_status,
        )


def main(path: str, apply: bool, allow_missing_sheets: bool = False) -> None:
    logger.info("Parsing FRU matrix workbook: {} (mode={})", path, "APPLY" if apply else "dry-run")
    parsed = parse_workbook(path, allow_missing_sheets=allow_missing_sheets)
    _report(parsed)
    links = parsed.links
    if not apply:
        logger.info("Dry run — nothing written. Re-run with --apply to upsert {} links.", len(links))
        return
    if parsed.unexpected_sheets:
        raise SystemExit(
            f"Refusing --apply: unexpected sheet(s) {parsed.unexpected_sheets} — map them in PARSERS "
            "or add them to KNOWN_SKIPPED_SHEETS first"
        )
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        inserted, updated = upsert_links(db, links)
        logger.info(
            "Upsert complete: {} inserted, {} updated, {} unchanged", inserted, updated, len(links) - inserted - updated
        )
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest the FRU_PN_TRAY matrix workbook into fru_links")
    parser.add_argument("xlsx", help="Path to the FRU matrix .xlsx workbook")
    parser.add_argument("--apply", action="store_true", help="Write to the database (default: dry run)")
    parser.add_argument(
        "--allow-missing-sheets",
        action="store_true",
        help="Tolerate mapped sheets missing from the workbook (deliberately partial workbooks only)",
    )
    args = parser.parse_args()
    main(args.xlsx, apply=args.apply, allow_missing_sheets=args.allow_missing_sheets)
