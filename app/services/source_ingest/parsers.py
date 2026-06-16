"""SP-Ingest parsers — read TRIO source files into raw (uncleaned) SourceRecords.

What: ``parse_inventory_sheet`` handles the operational inventory exports (Inventory 2.12.26 /
      Firesale / Foxconn) across .csv, .xlsx (streamed via openpyxl read-only) and the staged
      tab-delimited .txt captures — detecting the header row and the per-sheet column names.
      ``parse_sfdc_material_master`` STREAMS the multi-hundred-MB SFDC ``LSC1__Material__c`` CSV row by
      row (never loads the whole file), mapping the load-bearing columns + deep technical facets.
Called by: app/management/ingest_source_data.py (the parse stage). Both functions are
      generators — the CLI feeds their output through clean.py then consolidate.py.
Depends on: csv (stdlib), openpyxl (xlsx streaming), the SourceRecord dataclass + its
      SOURCE_KIND_* tags, and app.utils.normalization.normalize_quantity.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path

from loguru import logger

from app.services.source_ingest.models import (
    SOURCE_KIND_INVENTORY_SHEET,
    SOURCE_KIND_SFDC_MASTER,
    SourceRecord,
)
from app.utils.normalization import normalize_quantity

# ── Inventory-sheet column aliases ─────────────────────────────────────────
# Header detection: each logical field maps to the set of header labels seen across the
# operational sheets (lowercased, trimmed). "productdescription" (Inventory 2.12.26),
# "description" (Firesale/Foxconn), "productid" (Foxconn) all collapse to one field.
_SHEET_COLUMNS: dict[str, set[str]] = {
    "mpn": {"part number", "partnumber", "productid", "product id"},
    "description": {"productdescription", "product description", "description"},
    "condition": {"condition"},
    "category": {"commodity", "commodity code"},
    "quantity": {"on hand", "on-hand", "onhand", "qty avail", "qty", "quantity"},
}

# SFDC column → app spec_key for the deep facets the materials deep-filters consume. Only
# these are emitted; record_spec drops any key without a matching commodity schema, so a
# capacity on a non-drive card is harmlessly ignored downstream.
_SFDC_SPEC_COLUMNS: dict[str, str] = {
    "Capacity__c": "capacity_gb",
    "Legacy_RPM__c": "rpm",
    "Speed__c": "speed_mhz",
    "Form_Factor__c": "form_factor",
    "CPU__c": "cpu",
    "Pins__c": "pins",
    "Number_of_Cores__c": "number_of_cores",
    "Resolution__c": "resolution",
    "Connector_Type__c": "connector_type",
    "Backlight_Type__c": "backlight_type",
}

# SFDC description columns, in preference order (first non-empty wins).
_SFDC_DESCRIPTION_COLUMNS = (
    "Material_Description__c",
    "LSC1__Material_Detail_Description__c",
    "LSC1__Material_Short_Description__c",
    "LSC1__Common_Name__c",
)
# SFDC OEM/manufacturer NAME columns, in preference order. NOTE: per CATALOG.md's profile,
# LSC1__OEM__c / Brand__c are ~0% filled in this org — the real manufacturer signal is
# LSC1__Manufacturer_Brand__c, which holds Salesforce LOOKUP IDs that must resolve through
# LSC1__Manufacturers__c (parse_sfdc_manufacturers below); it is handled separately, never
# emitted verbatim.
_SFDC_OEM_COLUMNS = ("LSC1__OEM__c", "Brand__c")
_SFDC_MANUFACTURER_LOOKUP_COLUMN = "LSC1__Manufacturer_Brand__c"
# SFDC category columns, in preference order.
_SFDC_CATEGORY_COLUMNS = ("LSC1__Category__c", "Commodity_Code__c")


def _detect_text_encoding(path: Path) -> str:
    """Detect a source file's text encoding: UTF-8 (usual) or cp1252 fallback.

    SFDC/Excel exports are usually UTF-8 but sometimes carry stray cp1252 bytes (NBSP
    0xa0, smart quotes) that crash a strict UTF-8 stream mid-file. Streams the bytes
    through an incremental UTF-8 decoder; the first invalid byte switches the whole file
    to cp1252 (which decodes every byte, so parsing never crashes).
    """
    import codecs

    decoder = codecs.getincrementaldecoder("utf-8")()
    with open(path, "rb") as fh:
        while chunk := fh.read(1 << 20):
            try:
                decoder.decode(chunk)
            except UnicodeDecodeError:
                logger.info("_detect_text_encoding: {} is not valid UTF-8 — falling back to cp1252", path.name)
                return "cp1252"
    return "utf-8-sig"


def _first_nonempty(row: dict, columns: tuple[str, ...]) -> str | None:
    """Return the first non-empty/non-whitespace value across *columns* in *row*."""
    for col in columns:
        value = row.get(col)
        if value is None:
            continue
        stripped = str(value).strip()
        if stripped:
            return stripped
    return None


def _is_truthy(value) -> bool:
    """SFDC export booleans arrive as strings ("true"/"1"/"True").

    Coerce defensively.
    """
    if value is None:
        return False
    return str(value).strip().lower() in ("true", "1", "yes")


def _map_sheet_header(header: list[str]) -> dict[str, int]:
    """Map a detected header row to {logical_field: column_index}."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header):
        label = str(raw or "").strip().lower()
        for field_name, aliases in _SHEET_COLUMNS.items():
            if label in aliases and field_name not in mapping:
                mapping[field_name] = idx
    return mapping


def _looks_like_header(cells: list[str]) -> bool:
    """A row is the header iff it carries the MPN column AND a description column."""
    mapping = _map_sheet_header(cells)
    return "mpn" in mapping and "description" in mapping


def _row_to_record(cells: list[str], mapping: dict[str, int], source_file: str) -> SourceRecord | None:
    """Build a raw SourceRecord from a sheet data row using the detected column mapping.

    Returns None for rows too narrow to span the mapped columns (prose preamble /
    truncation notes in the staged .txt captures collapse to a single cell and are
    skipped here).
    """
    min_width = max(mapping.values()) + 1 if mapping else 0
    if len(cells) < min_width:
        return None

    def cell(field_name: str) -> str | None:
        idx = mapping.get(field_name)
        if idx is None or idx >= len(cells):
            return None
        value = cells[idx]
        if value is None:
            return None
        stripped = str(value).strip()
        return stripped or None

    raw_mpn = cell("mpn")
    if not raw_mpn:
        return None
    return SourceRecord(
        raw_mpn=raw_mpn,
        manufacturer=None,  # operational sheets embed OEM in the description (clean.py extracts)
        description=cell("description"),
        condition=cell("condition"),
        quantity=normalize_quantity(cell("quantity")),
        category=cell("category"),
        source_file=source_file,
        source_kind=SOURCE_KIND_INVENTORY_SHEET,
    )


def _iter_delimited_rows(path: Path) -> Iterator[list[str]]:
    """Yield rows from a .csv or staged .txt capture.

    .csv → ``csv.reader`` (RFC-4180 quoting, so commas inside quoted descriptions stay in one
    cell). The staged .txt captures are tab-delimited with a prose preamble, so we split each
    line on tab and let the header detector find the real header amid the noise (a non-tab
    prose line collapses to a single cell and is skipped by the header/row logic).
    """
    if path.suffix.lower() == ".txt":
        with open(path, encoding=_detect_text_encoding(path), newline="") as fh:
            for line in fh:
                line = line.rstrip("\n").rstrip("\r")
                if not line:
                    continue
                yield [c.strip() for c in line.split("\t")]
        return
    with open(path, encoding=_detect_text_encoding(path), newline="") as fh:
        for cells in csv.reader(fh):
            if not cells:
                continue
            yield [c.strip() for c in cells]


def _xlsx_cell_to_str(c) -> str:
    """Coerce an openpyxl cell value to its text form for the row pipeline.

    openpyxl returns numeric cells as int/float; a float-typed Part Number cell would
    str() to "5052089.0", whose dedup key (normalize_mpn_key strips the dot, KEEPING the
    trailing zero → "50520890") would silently fail to merge with the SFDC master's
    string MPN. Integral floats are therefore coerced through int() first. KNOWN
    LIMITATION (inherent to the source, not fixable here): a number-typed Excel cell has
    already lost its leading zeros inside the workbook ("005052089" stored as the number
    5052089) — only text-formatted MPN cells preserve them.
    """
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c)


def _iter_xlsx_rows(path: Path) -> Iterator[list[str]]:
    """Yield rows from an .xlsx via openpyxl read-only (streaming, never loads whole
    book)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield [_xlsx_cell_to_str(c) for c in row]
    finally:
        wb.close()


def parse_inventory_sheet(path: str | Path) -> Iterator[SourceRecord]:
    """Parse an operational inventory sheet into raw SourceRecords.

    Supports .csv (DictReader-grade), .xlsx (openpyxl read-only/streaming), and the
    staged tab-delimited .txt captures. The header row is auto-detected (the .txt
    captures carry a prose preamble before the real header), and columns are mapped per-
    sheet by their labels. Yields one raw (uncleaned) SourceRecord per data row;
    clean.py + consolidate.py finish it.
    """
    path = Path(path)
    source_file = path.name
    row_iter = _iter_xlsx_rows(path) if path.suffix.lower() in (".xlsx", ".xlsm") else _iter_delimited_rows(path)

    mapping: dict[str, int] | None = None
    yielded = 0
    for cells in row_iter:
        if mapping is None:
            if _looks_like_header(cells):
                mapping = _map_sheet_header(cells)
                logger.debug("parse_inventory_sheet: {} header detected: {}", source_file, mapping)
            continue
        rec = _row_to_record(cells, mapping, source_file)
        if rec is not None:
            yielded += 1
            yield rec
    if mapping is None:
        logger.warning("parse_inventory_sheet: no header row found in {}", source_file)
    else:
        logger.info("parse_inventory_sheet: {} → {} raw rows", source_file, yielded)


def parse_sfdc_manufacturers(path: str | Path) -> dict[str, str]:
    """Parse ``LSC1__Manufacturers__c.csv`` into a {salesforce_id: manufacturer_name}
    map.

    Resolves the ``LSC1__Manufacturer_Brand__c`` lookup IDs on the part master to real
    manufacturer names (CATALOG.md "manufacturer-lookup resolution"). Skips deleted rows
    and rows without a name.
    """
    lookup: dict[str, str] = {}
    with open(path, encoding=_detect_text_encoding(path), newline="") as fh:
        for row in csv.DictReader(fh):
            if _is_truthy(row.get("IsDeleted")):
                continue
            sfdc_id = (row.get("Id") or "").strip()
            name = (row.get("Name") or "").strip()
            if sfdc_id and name:
                lookup[sfdc_id] = name
    logger.info("parse_sfdc_manufacturers: {} → {} manufacturer names", Path(path).name, len(lookup))
    return lookup


def _resolve_sfdc_manufacturer(row: dict, manufacturer_lookup: dict[str, str] | None) -> str | None:
    """Resolve the manufacturer for one master row.

    Name columns (LSC1__OEM__c / Brand__c) win verbatim when filled; otherwise the
    ``LSC1__Manufacturer_Brand__c`` lookup ID resolves through *manufacturer_lookup*. An
    unresolvable lookup ID yields None — a raw Salesforce ID must NEVER be emitted as a
    manufacturer name.
    """
    name = _first_nonempty(row, _SFDC_OEM_COLUMNS)
    if name:
        return name
    lookup_id = (row.get(_SFDC_MANUFACTURER_LOOKUP_COLUMN) or "").strip()
    if lookup_id and manufacturer_lookup:
        return manufacturer_lookup.get(lookup_id)
    return None


def parse_sfdc_material_master(
    path: str | Path, manufacturer_lookup: dict[str, str] | None = None
) -> Iterator[SourceRecord]:
    """STREAM the SFDC ``LSC1__Material__c`` CSV part-master into raw SourceRecords.

    Uses csv.DictReader so the multi-hundred-MB file is read row-by-row (never wholly in memory).
    Maps MPN/OEM/description/category per CATALOG.md, emits only the non-empty deep facets,
    and skips rows whose ``IsDeleted`` is truthy. *manufacturer_lookup* (from
    parse_sfdc_manufacturers) resolves the ``LSC1__Manufacturer_Brand__c`` lookup IDs —
    without it those rows simply carry no manufacturer (an unresolved Salesforce ID is
    never emitted). Yields one raw SourceRecord per kept row.
    """
    path = Path(path)
    source_file = path.name
    kept = 0
    skipped_deleted = 0
    with open(path, encoding=_detect_text_encoding(path), newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if _is_truthy(row.get("IsDeleted")):
                skipped_deleted += 1
                continue
            raw_mpn = (row.get("LSC1__Material_Number__c") or "").strip()
            if not raw_mpn:
                continue
            specs: dict = {}
            for col, spec_key in _SFDC_SPEC_COLUMNS.items():
                value = row.get(col)
                if value is not None and str(value).strip():
                    specs[spec_key] = str(value).strip()
            kept += 1
            yield SourceRecord(
                raw_mpn=raw_mpn,
                manufacturer=_resolve_sfdc_manufacturer(row, manufacturer_lookup),
                description=_first_nonempty(row, _SFDC_DESCRIPTION_COLUMNS),
                condition=None,  # condition is per-unit, not on the master (CATALOG.md §1)
                quantity=normalize_quantity(row.get("LSC1__Total_Available_Inventory__c")),
                category=_first_nonempty(row, _SFDC_CATEGORY_COLUMNS),
                specs=specs,
                source_file=source_file,
                source_kind=SOURCE_KIND_SFDC_MASTER,
            )
    logger.info(
        "parse_sfdc_material_master: {} → {} rows ({} skipped IsDeleted)",
        source_file,
        kept,
        skipped_deleted,
    )
