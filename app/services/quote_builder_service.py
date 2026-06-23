# app/services/quote_builder_service.py
"""services/quote_builder_service.py — Business logic for the Quote Builder.

Loads requirement + offer data for the builder modal, applies smart defaults,
generates Excel exports. Decoupled from HTTP.

Called by: app.routers.quote_builder
Depends on: app.models (Requirement, Offer, Quote), openpyxl,
    app.routers.crm._helpers (_preload_last_quoted_prices — pricing history lookup)
"""

from __future__ import annotations

from sqlalchemy.orm import Session, joinedload

# Default thin-margin threshold for the Build-Quote guardrail. Matches the
# established floor used by proactive_min_margin_pct / buyplan_min_margin_pct (10%).
DEFAULT_MIN_MARGIN_PCT = 10.0

# Default markup applied to the best cost when no last-quoted price exists, used to
# pre-fill the sell-price field on the Build-Quote tab. Surfaced + editable in the UI.
DEFAULT_MARKUP_PCT = 20.0


def best_cost_for(db: Session, requirement_id: int) -> dict | None:
    """Best (lowest) unit cost across a requirement's ACTIVE offers — compute-on-read.

    Returns ``{"unit_cost": float, "offer_id": int}`` for the cheapest priced ACTIVE
    offer on *requirement_id*, or ``None`` when the requirement has no priced active
    offers. Mirrors the SHAPE of the resell ``ExcessLineItem.best_offer_unit_price``
    rollup (``excess_service.recompute_line_rollup``: min priced unit_price + the offer
    that provided it) but for the buyer-side ``Offer`` table, and uses the SAME
    ``OfferStatus.ACTIVE`` filter the quote builder applies in ``get_builder_data``.

    No schema/migration: this is a read-time query, not a stored column.
    """
    return best_costs_for(db, [requirement_id]).get(requirement_id)


def best_costs_for(db: Session, requirement_ids: list[int]) -> dict[int, dict]:
    """Batch best-cost rollup keyed by requirement id — one query, no N+1.

    For each id in *requirement_ids* that has at least one priced ACTIVE offer, the
    returned map carries ``{requirement_id: {"unit_cost": float, "offer_id": int}}``
    for the cheapest such offer. Requirements with no priced active offers are simply
    absent from the map (compute-on-read mirror of ``best_cost_for``).
    """
    if not requirement_ids:
        return {}

    from app.constants import OfferStatus
    from app.models import Offer

    rows = (
        db.query(Offer.requirement_id, Offer.id, Offer.unit_price)
        .filter(
            Offer.requirement_id.in_(requirement_ids),
            Offer.status == OfferStatus.ACTIVE,
            Offer.unit_price.isnot(None),
        )
        .all()
    )

    best: dict[int, dict] = {}
    for req_id, offer_id, unit_price in rows:
        cost = float(unit_price)
        current = best.get(req_id)
        if current is None or cost < current["unit_cost"]:
            best[req_id] = {"unit_cost": cost, "offer_id": offer_id}
    return best


def margin_guardrail(cost, sell, *, min_margin_pct: float = DEFAULT_MIN_MARGIN_PCT) -> str | None:
    """Short warning when a quote line is underwater or thin, else ``None`` (pure).

    Returns a customer-safe internal warning string when ``sell < cost`` (selling at a
    loss) or when ``margin% < min_margin_pct`` (thin margin), otherwise ``None``. A
    missing/zero sell price yields ``None`` — there is nothing to judge yet and we never
    divide by zero. Mirrors the buyplan/proactive "below threshold" check.
    """
    if cost is None or sell is None:
        return None
    cost_f = float(cost)
    sell_f = float(sell)
    if sell_f <= 0:
        return None
    if sell_f < cost_f:
        return f"Selling below cost (cost ${cost_f:,.4f} > sell ${sell_f:,.4f})"
    margin_pct = (sell_f - cost_f) / sell_f * 100
    if margin_pct < min_margin_pct:
        return f"Thin margin {margin_pct:.1f}% (below {min_margin_pct:.0f}% floor)"
    return None


def quote_export_context(quote) -> dict:
    """Build the CLEAN customer-facing export payload for *quote* (pure whitelist).

    The returned dict's ``lines`` carry ONLY ``part_number`` / ``manufacturer`` /
    ``quantity`` / ``condition`` / ``cost`` / ``sell`` / ``margin`` / ``extended`` — every
    vendor / offer / source identity field on the saved line (``vendor_name``,
    ``offer_id``, ``source``, ``material_card_id``, …) is STRIPPED here. The header carries
    only the quote number / revision / customer / date. Cleanliness is enforced at
    ASSEMBLY (the context explicitly enumerates the clean fields), not by hoping the
    template omits a leaky one — mirrors ``bid_back_service.bid_back_export_context``.
    This is the single source ``generate_quote_report_pdf`` renders ``quote_report.html``
    from.
    """
    from datetime import date

    lines: list[dict] = []
    subtotal = 0.0
    for li in quote.line_items or []:
        qty = li.get("qty") or 0
        cost = float(li.get("cost_price") or 0)
        sell = float(li.get("sell_price") or 0)
        margin_pct = li.get("margin_pct")
        margin = float(margin_pct) if margin_pct is not None else (((sell - cost) / sell * 100) if sell else 0.0)
        extended = round(sell * qty, 2)
        subtotal += extended
        # WHITELIST — explicitly enumerate the clean fields. No vendor/offer/source key
        # on the saved line is referenced, so nothing leaky can ride along.
        lines.append(
            {
                "part_number": li.get("mpn"),
                "manufacturer": li.get("manufacturer"),
                "quantity": qty,
                "condition": li.get("condition"),
                "cost": cost,
                "sell": sell,
                "margin": round(margin, 2),
                "extended": extended,
            }
        )

    customer = None
    if quote.customer_site_id is not None:
        site = getattr(quote, "customer_site", None)
        if site is not None:
            customer = getattr(site, "site_name", None)

    return {
        "quote_number": quote.quote_number,
        "revision": quote.revision or 1,
        "customer": customer,
        "date": date.today().isoformat(),
        "lines": lines,
        "subtotal": round(subtotal, 2),
        "line_count": len(lines),
    }


def get_builder_data(
    req_id: int,
    db: Session,
    requirement_ids: list[int] | None = None,
) -> list[dict]:
    """Load requirements + offers for the quote builder modal."""
    from app.constants import OfferStatus
    from app.models import Requirement

    query = db.query(Requirement).options(joinedload(Requirement.offers)).filter(Requirement.requisition_id == req_id)
    if requirement_ids:
        query = query.filter(Requirement.id.in_(requirement_ids))
    query = query.order_by(Requirement.id)
    requirements = query.all()

    lines = []
    for r in requirements:
        offers_data = [
            {
                "id": o.id,
                "vendor_name": o.vendor_name,
                "unit_price": float(o.unit_price) if o.unit_price else 0,
                "qty_available": o.qty_available or 0,
                "lead_time": o.lead_time,
                "date_code": o.date_code,
                "condition": o.condition,
                "packaging": o.packaging,
                "moq": o.moq,
                "confidence": o.parse_confidence,
                "material_card_id": o.material_card_id,
                "notes": o.notes,
            }
            for o in r.offers
            if o.status == OfferStatus.ACTIVE
        ]

        lines.append(
            {
                "requirement_id": r.id,
                "mpn": r.primary_mpn,
                "manufacturer": r.manufacturer,
                "target_qty": r.target_qty or 0,
                "target_price": float(r.target_price) if r.target_price else None,
                "customer_pn": r.customer_pn,
                "date_codes": r.date_codes,
                "condition": r.condition,
                "packaging": r.packaging,
                "firmware": r.firmware,
                "hardware_codes": r.hardware_codes,
                "sale_notes": r.sale_notes,
                "need_by_date": str(r.need_by_date) if r.need_by_date else None,
                "offers": offers_data,
                "offer_count": len(offers_data),
                "status": "unknown",
                "selected_offer_id": None,
                "sell_price": None,
                "sell_price_manual": False,
                "buyer_notes": "",
                "pricing_history": None,
            }
        )

    # Load pricing history for all MPNs in one pass
    try:
        from app.routers.crm._helpers import _preload_last_quoted_prices

        quoted_prices = _preload_last_quoted_prices(db)
        for line in lines:
            mpn_key = (line["mpn"] or "").upper().strip()
            lq = quoted_prices.get(mpn_key)
            if lq:
                line["pricing_history"] = {
                    "avg_price": lq.get("sell_price"),  # Most recent sell price, not a true average
                    "price_range": None,
                    "recent": [
                        {
                            "quote_number": lq.get("quote_number", ""),
                            "date": lq.get("date", ""),
                            "cost": lq.get("cost_price", lq.get("sell_price")),
                            "sell": lq.get("sell_price"),
                            "margin": lq.get("margin_pct"),
                            "result": lq.get("result"),
                        }
                    ],
                }
    except Exception as e:
        from loguru import logger

        logger.warning("Pricing history unavailable for req {}: {} — lines will show no history", req_id, e)

    return lines


def build_quote_tab_data(
    db: Session,
    req_id: int,
    *,
    markup_pct: float = DEFAULT_MARKUP_PCT,
) -> list[dict]:
    """Per-line data for the in-workspace Build-Quote tab (single-stage inline
    assembly).

    Each line carries the best-cost reference (``best_costs_for``), its ACTIVE offers (the
    cheapest flagged ``is_best``), and a seeded ``sell_price``: the last-quoted price
    (``_preload_last_quoted_prices``) when known, else best-cost × (1 + markup). The
    template/Alpine layer reads these straight through — the seed is the only smart default.

    Mirrors the SHAPE of ``resell._build_bid_context`` line items (a best reference + a
    pre-filled "our price"), but for the buyer-side ``Requirement``/``Offer`` tables.
    """
    from app.constants import OfferStatus
    from app.models import Requirement

    requirements = (
        db.query(Requirement)
        .options(joinedload(Requirement.offers))
        .filter(Requirement.requisition_id == req_id)
        .order_by(Requirement.id)
        .all()
    )

    best_costs = best_costs_for(db, [r.id for r in requirements])

    try:
        from app.routers.crm._helpers import _preload_last_quoted_prices

        last_quoted = _preload_last_quoted_prices(db)
    except Exception as e:  # pragma: no cover - history is best-effort
        from loguru import logger

        logger.warning("Pricing history unavailable for req {}: {} — seeds fall back to markup", req_id, e)
        last_quoted = {}

    markup_factor = 1 + (markup_pct / 100.0)

    lines: list[dict] = []
    for r in requirements:
        best = best_costs.get(r.id)
        best_cost = best["unit_cost"] if best else None
        best_offer_id = best["offer_id"] if best else None

        offers = [
            {
                "id": o.id,
                "vendor_name": o.vendor_name,
                "unit_price": float(o.unit_price) if o.unit_price else 0.0,
                "qty_available": o.qty_available or 0,
                "lead_time": o.lead_time,
                "date_code": o.date_code,
                "condition": o.condition,
                "packaging": o.packaging,
                "moq": o.moq,
                "material_card_id": o.material_card_id,
                "is_best": o.id == best_offer_id,
            }
            for o in r.offers
            if o.status == OfferStatus.ACTIVE and o.unit_price is not None
        ]
        offers.sort(key=lambda o: o["unit_price"])

        # Seed the sell price: last-quoted wins, else best-cost × markup.
        lq = last_quoted.get((r.primary_mpn or "").upper().strip())
        seed = lq.get("sell_price") if lq else None
        seed_source = "last_quoted" if seed is not None else None
        if seed is None and best_cost is not None:
            seed = round(best_cost * markup_factor, 4)
            seed_source = "markup"

        lines.append(
            {
                "requirement_id": r.id,
                "mpn": r.primary_mpn,
                "manufacturer": r.manufacturer,
                "qty": r.target_qty or 0,
                "condition": r.condition,
                "best_cost": best_cost,
                "best_offer_id": best_offer_id,
                "offers": offers,
                "offer_count": len(offers),
                "sell_seed": float(seed) if seed is not None else None,
                "seed_source": seed_source,
            }
        )

    return lines


def apply_smart_defaults(lines: list[dict]) -> None:
    """Apply smart defaults to builder lines in-place.

    - 1 offer: auto-select, status = decided
    - Multiple offers: status = needs_review
    - 0 offers: status = no_offers
    """
    for line in lines:
        offers = line.get("offers", [])
        if len(offers) == 1:
            line["status"] = "decided"
            line["selected_offer_id"] = offers[0]["id"]
            line["sell_price"] = offers[0]["unit_price"]
            line["sell_price_manual"] = False
        elif len(offers) > 1:
            line["status"] = "needs_review"
            line["selected_offer_id"] = None
        else:
            line["status"] = "no_offers"
            line["selected_offer_id"] = None


def build_excel_export(
    line_items: list[dict],
    quote_number: str,
    customer_name: str,
) -> bytes:
    """Generate a styled Excel workbook from quote line items.

    Returns raw bytes.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = quote_number or "Quote"

    header_fill = PatternFill(start_color="3D6895", end_color="3D6895", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    columns = [
        ("MPN", 20),
        ("Manufacturer", 18),
        ("Qty", 10),
        ("Unit Price", 14),
        ("Extended Price", 16),
        ("Lead Time", 14),
        ("Date Codes", 14),
        ("Condition", 12),
        ("Packaging", 14),
        ("MOQ", 10),
        ("Vendor", 20),
    ]

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, item in enumerate(line_items, 2):
        qty = item.get("qty") or 0
        sell = item.get("sell_price") or 0
        ws.cell(row=row_idx, column=1, value=item.get("mpn", ""))
        ws.cell(row=row_idx, column=2, value=item.get("manufacturer", ""))
        ws.cell(row=row_idx, column=3, value=qty)
        price_cell = ws.cell(row=row_idx, column=4, value=sell)
        price_cell.number_format = "$#,##0.0000"
        ext_cell = ws.cell(row=row_idx, column=5, value=round(qty * sell, 2))
        ext_cell.number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=6, value=item.get("lead_time", ""))
        ws.cell(row=row_idx, column=7, value=item.get("date_code", ""))
        ws.cell(row=row_idx, column=8, value=item.get("condition", ""))
        ws.cell(row=row_idx, column=9, value=item.get("packaging", ""))
        ws.cell(row=row_idx, column=10, value=item.get("moq", ""))
        ws.cell(row=row_idx, column=11, value=item.get("vendor_name", ""))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_quote_from_builder(
    db: Session,
    req_id: int,
    payload,
    user,
) -> dict:
    """Create or revise a Quote from the builder's save payload.

    If payload.quote_id is set, marks the old quote as 'revised' and creates a new
    revision with the updated line items. Otherwise creates a fresh quote.

    Fires the same hooks as create_quote: requisition state transition,
    per-requirement sourcing status update, and knowledge ledger capture.
    """
    from loguru import logger

    from app.constants import QuoteStatus
    from app.models import Quote, QuoteLine, Requisition
    from app.services.crm_service import next_quote_number

    req = db.get(Requisition, req_id)
    if not req:
        raise ValueError("Requisition not found")

    line_items = [
        {
            "mpn": li.mpn,
            "manufacturer": li.manufacturer,
            "qty": li.qty,
            "cost_price": li.cost_price,
            "sell_price": li.sell_price,
            "margin_pct": li.margin_pct,
            "lead_time": li.lead_time,
            "date_code": li.date_code,
            "condition": li.condition,
            "packaging": li.packaging,
            "moq": li.moq,
            "offer_id": li.offer_id,
            "material_card_id": li.material_card_id,
            "notes": li.notes,
        }
        for li in payload.lines
    ]

    total_sell = sum(li["qty"] * li["sell_price"] for li in line_items)
    total_cost = sum(li["qty"] * li["cost_price"] for li in line_items)
    margin_pct = round((total_sell - total_cost) / total_sell * 100, 2) if total_sell > 0 else 0

    # Rename old quote to Q-XXXX-R{n} so the new revision keeps the canonical number
    revision = 1
    old_quote = db.get(Quote, payload.quote_id) if payload.quote_id else None
    if old_quote:
        old_revision = old_quote.revision or 1
        quote_number = old_quote.quote_number
        revision = old_revision + 1
        old_quote.quote_number = f"{quote_number}-R{old_revision}"
        old_quote.status = QuoteStatus.REVISED
    else:
        quote_number = next_quote_number(db)

    # Advance requisition status to "quoting" if appropriate
    from app.constants import RequisitionStatus

    if req.status in (RequisitionStatus.ACTIVE, RequisitionStatus.SOURCING, RequisitionStatus.OFFERS):
        try:
            from app.services.requisition_state import transition as req_transition

            req_transition(req, "quoting", user, db)
        except ValueError:
            pass  # already in quoting or later state

    # Load customer site for default payment/shipping terms
    from app.models import CustomerSite

    site = db.get(CustomerSite, req.customer_site_id) if req.customer_site_id else None

    quote = Quote(
        requisition_id=req_id,
        customer_site_id=req.customer_site_id,
        quote_number=quote_number,
        revision=revision,
        line_items=line_items,
        subtotal=total_sell,
        total_cost=total_cost,
        total_margin_pct=margin_pct,
        payment_terms=payload.payment_terms or (site.payment_terms if site else None),
        shipping_terms=payload.shipping_terms or (site.shipping_terms if site else None),
        validity_days=payload.validity_days,
        notes=payload.notes,
        created_by_id=user.id,
    )
    db.add(quote)
    db.flush()

    for li in line_items:
        db.add(
            QuoteLine(
                quote_id=quote.id,
                material_card_id=li.get("material_card_id"),
                offer_id=li.get("offer_id"),
                mpn=li.get("mpn", ""),
                manufacturer=li.get("manufacturer"),
                qty=li.get("qty"),
                cost_price=li.get("cost_price"),
                sell_price=li.get("sell_price"),
                margin_pct=li.get("margin_pct"),
                currency="USD",
            )
        )

    db.commit()

    # Fire per-requirement sourcing status hook (same as create_quote)
    try:
        from app.services.requirement_status import on_quote_built

        offer_ids = [li.get("offer_id") for li in line_items if li.get("offer_id")]
        if offer_ids:
            from app.models import Offer as OfferModel

            offers_used = db.query(OfferModel).filter(OfferModel.id.in_(offer_ids)).all()
            requirement_ids = list({o.requirement_id for o in offers_used if o.requirement_id})
            if requirement_ids:
                on_quote_built(requirement_ids, db, actor=user)
                db.commit()
    except Exception as e:
        logger.warning("Requirement status update (on_quote_built) failed: {}", e)

    # Knowledge ledger capture (same as create_quote)
    try:
        from app.services.knowledge_service import capture_quote_fact

        capture_quote_fact(db, quote=quote, user_id=user.id)
    except Exception as e:
        logger.warning("Knowledge auto-capture (quote) failed: {}", e)

    return {
        "ok": True,
        "quote_id": quote.id,
        "quote_number": quote.quote_number,
        "revision": quote.revision,
    }
