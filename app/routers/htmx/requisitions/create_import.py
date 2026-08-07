"""Unified create/import modal — forms, AI parse/save, customer typeahead/lookup/quick-
create.

W4.8 split of the 1,473-line app/routers/htmx/requisitions.py — pure structural
move: URLs and behavior unchanged; every route attaches to the shared router
imported from .common (registration assembled in __init__).
"""

import asyncio
import html as html_mod
import json

from fastapi import Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse
from loguru import logger
from sqlalchemy.orm import Session, selectinload

from ....constants import (
    OfferCondition,
    RequisitionStatus,
)
from ....database import get_db
from ....dependencies import require_user
from ....models import (
    Company,
    CustomerSite,
    Requisition,
    User,
)
from ....services.freeform_parser_service import parse_freeform_rfq
from ....template_env import template_response
from ....utils.claude_client import claude_configured
from ....utils.sql_helpers import escape_like
from .._shared import _base_ctx
from .common import router

# Import-parse upload cap (P2.6) — same 10MB convention as resell.py's
# MAX_UPLOAD_BYTES / requisitions/requirements.py's inline 10_000_000 check.
MAX_IMPORT_UPLOAD_BYTES = 10 * 1024 * 1024


def _parse_xlsx_rows(content: bytes) -> str:
    """Parse an uploaded XLSX/XLS workbook into tab-separated text.

    Sync (openpyxl has no async API) — always dispatched via ``asyncio.to_thread``
    from ``requisition_import_parse`` (P2.6), since ``load_workbook`` + full-sheet
    iteration can block the event loop for a large workbook.
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
    return "\n".join(rows)


@router.get("/v2/partials/requisitions/create-form", response_class=HTMLResponse)
async def requisition_create_form(
    request: Request,
    prospect_id: str = "",
    company_id: str = "",
    customer_name: str = "",
    user: User = Depends(require_user),
    db: Session = Depends(get_db),
):
    """Return the create requisition modal form.

    Optionally PREFILLED when launched from a claimed prospect's "Create Requisition"
    button (H1): ``company_id`` resolves the customer's HQ site so the picker opens with
    it selected, and ``prospect_id`` rides through as a hidden field so a successful save
    flips the prospect to CONVERTED (see requisition_import_save). With no params it is
    the plain create modal.
    """
    ctx = _base_ctx(request, user, "requisitions")

    prefill_site_id = ""
    prefill_customer_name = ""
    if company_id.strip().isdigit():
        company = db.get(Company, int(company_id))
        if company:
            prefill_customer_name = company.name
            site = (
                db.query(CustomerSite)
                .filter(CustomerSite.company_id == company.id, CustomerSite.is_active.is_(True))
                .order_by(CustomerSite.id)
                .first()
            )
            if site:
                prefill_site_id = str(site.id)
                prefill_customer_name = f"{company.name} — {site.site_name}"
    # Fall back to the passed-in customer name (e.g. the prospect name) when the company
    # has no match/site — the picker still shows the name so the buyer isn't lost.
    if not prefill_customer_name and customer_name.strip():
        prefill_customer_name = customer_name.strip()

    ctx.update(
        {
            "prefill_prospect_id": prospect_id.strip(),
            "prefill_customer_site_id": prefill_site_id,
            "prefill_customer_name": prefill_customer_name,
            "prefill_req_name": f"{prefill_customer_name} RFQ" if prefill_customer_name else "",
        }
    )
    return template_response("htmx/partials/requisitions/unified_modal.html", ctx)


@router.get("/v2/partials/requisitions/import-form", response_class=HTMLResponse)
async def requisition_import_form(
    request: Request,
    user: User = Depends(require_user),
):
    """Return the import requisition modal form."""
    ctx = _base_ctx(request, user, "requisitions")
    return template_response("htmx/partials/requisitions/unified_modal.html", ctx)


@router.post("/v2/partials/requisitions/import-parse", response_class=HTMLResponse)
async def requisition_import_parse(
    request: Request,
    name: str = Form(...),
    customer_name: str = Form(""),
    customer_site_id: str = Form(""),
    deadline: str = Form(""),
    urgency: str = Form("normal"),
    raw_text: str = Form(""),
    file: UploadFile | None = File(None),
    user: User = Depends(require_user),
):
    """Parse pasted text or uploaded file with AI, return editable preview."""
    json_mode = request.query_params.get("format") == "json"

    # Keys-off honesty (spec §7): with no AI key the parse cannot run — say so
    # instead of 500ing inside claude_client (ClaudeUnavailableError). Manual row
    # entry in the modal keeps working; only the bulk AI fill is off.
    if not claude_configured():
        message = "AI is off — enter lines or paste when enabled"
        if json_mode:
            from fastapi.responses import JSONResponse

            return JSONResponse({"error": message, "requirements": []})
        return template_response(
            "htmx/partials/shared/_ai_off_banner.html",
            {"request": request, "message": message},
        )

    # Extract text from file if uploaded
    text = raw_text.strip()
    if file and file.filename:
        content = await file.read()
        if len(content) > MAX_IMPORT_UPLOAD_BYTES:
            req_id = getattr(request.state, "request_id", "unknown")
            message = "File too large — 10MB maximum."
            if json_mode:
                from fastapi.responses import JSONResponse

                return JSONResponse(
                    status_code=413,
                    content={"error": message, "status_code": 413, "request_id": req_id},
                )
            return HTMLResponse(
                f'<div class="p-4 text-center text-sm text-rose-600 bg-rose-50 rounded-lg border border-rose-200">'
                f"{message}"
                "</div>",
                status_code=413,
            )
        fname = file.filename.lower()
        if fname.endswith((".xlsx", ".xls")):
            # openpyxl has no async API and full-sheet iteration can block the event
            # loop for a large workbook — parse on a worker thread (P2.6).
            text = await asyncio.to_thread(_parse_xlsx_rows, content)
        elif fname.endswith(".csv"):
            text = content.decode("utf-8", errors="replace")
        else:
            text = content.decode("utf-8", errors="replace")

    if not text:
        if json_mode:
            from fastapi.responses import JSONResponse

            return JSONResponse({"error": "No data provided", "requirements": []})
        return HTMLResponse(
            '<div class="p-4 text-center text-sm text-rose-600 bg-rose-50 rounded-lg border border-rose-200">'
            "No data provided. Paste text or upload a file."
            "</div>"
        )

    # AI parse
    result = await parse_freeform_rfq(text)
    requirements = result.get("requirements", []) if result else []

    # Use AI-extracted name/customer as fallback if user left them blank
    if not name.strip() and result:
        name = result.get("name", "Untitled")
    if not customer_name.strip() and result:
        customer_name = result.get("customer_name", "")

    if json_mode:
        from fastapi.responses import JSONResponse

        return JSONResponse(
            {
                "requirements": requirements,
                "inferred_name": name,
                "inferred_customer": customer_name,
            }
        )

    ctx = _base_ctx(request, user, "requisitions")
    ctx.update(
        {
            "requirements": requirements,
            "req_name": name,
            "customer_name": customer_name,
            "customer_site_id": customer_site_id,
            "deadline": deadline,
            "urgency": urgency,
            "count": len(requirements),
        }
    )
    return template_response("htmx/partials/requisitions/unified_modal.html", ctx)


@router.post("/v2/partials/requisitions/import-save", response_class=HTMLResponse)
async def requisition_import_save(
    request: Request,
    name: str = Form(...),
    customer_name: str = Form(""),
    customer_site_id: str = Form(""),
    deadline: str = Form(""),
    urgency: str = Form("normal"),
    prospect_id: str = Form(""),
    hotlist: bool = Form(False),
    user: User = Depends(require_user),
    db: Session = Depends(get_db),
):
    """Save AI-parsed requirements as a new requisition.

    When ``prospect_id`` is present (the modal was launched from a claimed prospect's
    "Create Requisition" button), the newly-created requisition flips that prospect to
    CONVERTED — closing the prospect→opportunity loop (H1/M4).

    When ``hotlist`` is set, the requisition is created as a monitored Hot List
    (``RequisitionStatus.HOTLIST``) rather than an active sourcing deal (OPEN): parts are
    stored + market data built, and the Proactive matcher surfaces offers when stock
    appears — nothing is sourced. That matcher joins Company on ``Requisition.company_id``,
    so ``company_id`` is populated from the chosen site for every create (hotlist or not),
    guarded for the no-site case.
    """
    from app.utils.normalization import parse_substitute_mpns

    form = await request.form()

    # Collect requirement rows from indexed form fields
    requirements = []
    idx = 0
    while f"reqs[{idx}].primary_mpn" in form:
        mpn = form.get(f"reqs[{idx}].primary_mpn", "").strip()
        if mpn:
            # Prefer the structured substitutes_json (mpn + manufacturer per sub) the modal
            # posts; fall back to the legacy comma-joined MPN string. parse_substitute_mpns()
            # normalizes either into the canonical [{"mpn", "manufacturer"}] list format
            # (CLAUDE.md "Substitutes Format") — the raw string list was the legacy bug.
            subs_input: list = []
            subs_json_raw = form.get(f"reqs[{idx}].substitutes_json", "").strip()
            if subs_json_raw:
                try:
                    parsed = json.loads(subs_json_raw)
                    if isinstance(parsed, list):
                        subs_input = parsed
                except (ValueError, TypeError):
                    subs_input = []
            if not subs_input:
                subs_input = [s.strip() for s in form.get(f"reqs[{idx}].substitutes", "").split(",") if s.strip()]
            requirements.append(
                {
                    "primary_mpn": mpn,
                    "target_qty": int(form.get(f"reqs[{idx}].target_qty", "1") or "1"),
                    "brand": form.get(f"reqs[{idx}].brand", "").strip() or None,
                    "target_price": float(form.get(f"reqs[{idx}].target_price") or "0") or None,
                    "condition": form.get(f"reqs[{idx}].condition", OfferCondition.NEW).strip(),
                    "customer_pn": form.get(f"reqs[{idx}].customer_pn", "").strip() or None,
                    "date_codes": form.get(f"reqs[{idx}].date_codes", "").strip() or None,
                    "packaging": form.get(f"reqs[{idx}].packaging", "").strip() or None,
                    "manufacturer": form.get(f"reqs[{idx}].manufacturer", "").strip(),
                    "substitutes": parse_substitute_mpns(subs_input, mpn),
                    "firmware": form.get(f"reqs[{idx}].firmware", "").strip() or None,
                    "hardware_codes": form.get(f"reqs[{idx}].hardware_codes", "").strip() or None,
                    "description": form.get(f"reqs[{idx}].description", "").strip() or None,
                    "package_type": form.get(f"reqs[{idx}].package_type", "").strip() or None,
                    "revision": form.get(f"reqs[{idx}].revision", "").strip() or None,
                    "need_by_date": form.get(f"reqs[{idx}].need_by_date", "").strip() or None,
                    "sale_notes": form.get(f"reqs[{idx}].sale_notes", "").strip() or None,
                }
            )
        idx += 1

    if not requirements:
        return HTMLResponse(
            '<div class="p-4 text-center text-sm text-rose-600 bg-rose-50 rounded-lg border border-rose-200">'
            "No valid parts to save."
            "</div>"
        )

    # Create requisition. Populate company_id from the chosen site so the Proactive
    # matcher's Company join resolves (a Hot List req with no company_id gets zero matches).
    site_id = int(customer_site_id) if customer_site_id.strip() else None
    company_id = None
    if site_id is not None:
        site = db.get(CustomerSite, site_id)
        if site is not None:
            company_id = site.company_id
    req = Requisition(
        name=name.strip() or "Untitled",
        customer_name=customer_name.strip() or None,
        customer_site_id=site_id,
        company_id=company_id,
        deadline=deadline.strip() or None,
        urgency=urgency,
        status=RequisitionStatus.HOTLIST if hotlist else RequisitionStatus.OPEN,
        created_by=user.id,
        claimed_by_id=user.id,
    )
    db.add(req)
    db.flush()

    # Create requirements through THE pipeline (services/requirement_service.py, spec §9):
    # normalization, material cards, tag propagation, task auto-gen, dup detection,
    # datasheet capture — one implementation for every creation surface.
    from ....services.requirement_service import create_requirements_ui

    result = await create_requirements_ui(db, req, requirements, actor_id=user.id)
    added = len(result.created)
    db.commit()

    # Prospect → opportunity handoff (H1/M4): if this modal was launched from a claimed
    # prospect, flip it to CONVERTED. Best-effort — the requisition is already committed,
    # so a conversion hiccup must never fail the save.
    if prospect_id.strip().isdigit():
        from ....services.prospect_claim import mark_prospect_converted

        try:
            mark_prospect_converted(int(prospect_id), user.id, db)
        except Exception:
            logger.warning("Prospect {} conversion after requisition create failed", prospect_id, exc_info=True)

    # Return success — close modal + toast, and fire reqListRefresh so whichever surface
    # opened this modal refreshes itself. The old snippet hard-targeted #parts-list, which
    # exists only in the parts workspace — opened from the requisitions list it hit
    # htmx:targetError and nothing refreshed. Both surfaces now listen for
    # `reqListRefresh from:body` (parts/workspace.html #parts-list, list.html hidden hook).
    safe_added = int(added)  # safe: server-computed int
    toast = f"Requisition created with {safe_added} parts"
    if result.duplicates:
        # UI dup detection (spec §9): same part quoted to this customer site in the
        # last 30 days — informational, the save already happened.
        toast += f" — {int(len(result.duplicates))} possible duplicate(s) in the last 30 days"
    resp = HTMLResponse(
        "<script>"
        "window.dispatchEvent(new CustomEvent('close-modal'));"
        f"Alpine.store('toast').message = '{toast}';"
        "Alpine.store('toast').type = 'success';"
        "Alpine.store('toast').show = true;"
        "</script>"
    )
    resp.headers["HX-Trigger"] = "reqListRefresh"
    return resp


@router.get("/v2/partials/requisitions/customer-typeahead", response_class=HTMLResponse)
async def customers_typeahead_dropdown(
    request: Request,
    q: str = "",
    user: User = Depends(require_user),
    db: Session = Depends(get_db),
):
    """P5.2: server-rendered debounced dropdown for the unified requisition modal's
    customer picker (unified_modal.html, customerPicker() in htmx_app.js).

    Runs the same active-Company + site query the retired JSON
    `/api/companies/typeahead` endpoint (crm.companies.companies_typeahead, removed —
    it had no remaining consumers once this HTML sibling replaced its only caller)
    used to serve, filtered server-side by `q` so the picker is a real hx-get swap
    instead of a client-side fetch-all + filter.
    """
    query = q.strip()
    companies_q = db.query(Company).filter(Company.is_active.is_(True)).options(selectinload(Company.sites))
    if query:
        companies_q = companies_q.filter(Company.name.ilike(f"%{escape_like(query)}%", escape="\\"))
    companies = companies_q.order_by(Company.name).limit(20).all()
    ctx = {
        "request": request,
        "companies": [{"id": c.id, "name": c.name, "sites": [s for s in c.sites if s.is_active]} for c in companies],
    }
    return template_response("htmx/partials/requisitions/_customer_typeahead_results.html", ctx)


@router.post("/v2/partials/customers/lookup", response_class=HTMLResponse)
async def customer_lookup(
    request: Request,
    company_name: str = Form(...),
    location: str = Form(""),
    user: User = Depends(require_user),
):
    """AI-powered company lookup using Claude with web search."""
    from app.utils.claude_client import claude_json
    from app.utils.claude_errors import ClaudeError, ClaudeUnavailableError

    search_query = company_name.strip()
    if location.strip():
        search_query += f", {location.strip()}"

    try:
        result = await claude_json(
            prompt=f"Search the web for this company: {search_query}\n\n"
            f"Find their official website, main phone number, and physical address.\n\n"
            f"Return ONLY a JSON object with these fields:\n"
            f'{{"company_name": "...", "website": "...", "phone": "...", '
            f'"address_line1": "...", "city": "...", "state": "...", "zip": "...", "country": "..."}}\n\n'
            f"Use empty strings for any field you cannot verify from search results. "
            f"Do NOT guess or make up information — only include data you found online.",
            system="You look up company information using web search. "
            "ONLY return data you can verify from search results. "
            "If you cannot find a phone number or address, return empty strings — never guess.",
            model_tier="smart",
            max_tokens=512,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
            timeout=45,
        )
    except (ClaudeUnavailableError, ClaudeError):
        result = None

    if not result:
        return HTMLResponse(
            '<p class="text-xs text-rose-500 mt-1">Could not look up company. Enter details manually.</p>'
        )

    # Render an approval card — escape all AI-provided strings for XSS safety
    # html_mod.escape() for HTML display context
    name = html_mod.escape(result.get("company_name", company_name))
    website = html_mod.escape(result.get("website", ""))
    phone = html_mod.escape(result.get("phone", ""))
    addr_parts = [
        p
        for p in [
            result.get("address_line1", ""),
            result.get("city", ""),
            (result.get("state", "") + " " + result.get("zip", "")).strip(),
            result.get("country", ""),
        ]
        if p
    ]
    address_display = html_mod.escape(", ".join(addr_parts))

    # json.dumps() for values embedded in JavaScript — handles quotes,
    # backslashes, </script> injection, etc.  Produces a quoted string
    # like "O\u0027Brien Corp" that is safe inside JS.
    name_js = json.dumps(result.get("company_name", company_name))
    website_js = json.dumps(result.get("website", ""))
    phone_js = json.dumps(result.get("phone", ""))
    addr1_js = json.dumps(result.get("address_line1", ""))
    city_js = json.dumps(result.get("city", ""))
    state_js = json.dumps(result.get("state", ""))
    zip_js = json.dumps(result.get("zip", ""))
    country_js = json.dumps(result.get("country", "US"))

    html_out = f"""
    <div class="mt-2 p-3 bg-emerald-50 border border-emerald-200 rounded-lg text-xs space-y-1">
      <div class="flex items-center justify-between">
        <span class="font-semibold text-emerald-700">Found: {name}</span>
      </div>
      {"<div class='text-gray-600'>🌐 " + website + "</div>" if website else ""}
      {"<div class='text-gray-600'>📞 " + phone + "</div>" if phone else ""}
      {"<div class='text-gray-600'>📍 " + address_display + "</div>" if address_display else ""}
      <div class="flex gap-2 mt-2">
        <button type="button" onclick="(async function(btn){{
            btn.disabled=true; btn.textContent='Saving...';
            var fd=new FormData();
            fd.append('company_name',{name_js});
            fd.append('website',{website_js});
            fd.append('phone',{phone_js});
            fd.append('address_line1',{addr1_js});
            fd.append('city',{city_js});
            fd.append('state',{state_js});
            fd.append('zip',{zip_js});
            fd.append('country',{country_js});
            try{{
              var r=await fetch('/v2/partials/customers/quick-create',{{method:'POST',body:fd}});
              var html=await r.text();
              var el=btn.closest('.space-y-1');
              el.replaceChildren();
              el.insertAdjacentHTML('afterbegin',html);
              var meta=el.querySelector('[data-site-id]');
              if(meta)document.dispatchEvent(new CustomEvent('customer-created',{{
                detail:{{siteId:meta.dataset.siteId,displayName:meta.dataset.display}}
              }}));
            }}catch(e){{console.error('quick-create failed:',e);btn.textContent='Failed — retry';btn.disabled=false;}}
          }})(this)"
                class="px-3 py-1 text-xs font-semibold bg-emerald-600 text-white rounded hover:bg-emerald-700">
          Use This Customer
        </button>
      </div>
    </div>
    """
    return HTMLResponse(html_out)


@router.post("/v2/partials/customers/quick-create", response_class=HTMLResponse)
async def customer_quick_create(
    request: Request,
    company_name: str = Form(...),
    website: str = Form(""),
    phone: str = Form(""),
    address_line1: str = Form(""),
    city: str = Form(""),
    state: str = Form(""),
    zip: str = Form(""),
    country: str = Form("US"),
    user: User = Depends(require_user),
    db: Session = Depends(get_db),
):
    """Create Company + default site from AI lookup, return JS to select it in
    picker."""
    from app.cache.decorators import invalidate_prefix

    # Check for duplicates
    existing = db.query(Company).filter(Company.name.ilike(escape_like(company_name.strip()), escape="\\")).first()
    if existing:
        site = existing.sites[0] if existing.sites else None
        site_id = site.id if site else ""
        display = html_mod.escape(f"{existing.name} — {site.site_name}" if site else existing.name)
        return HTMLResponse(
            f'<div class="mt-1 p-2 bg-amber-50 border border-amber-200 rounded text-xs text-amber-700">'
            f"Customer already exists. Selected automatically."
            f"</div>"
            f'<span class="hidden" data-site-id="{site_id}" data-display="{display}"></span>'
        )

    # Create company
    domain = ""
    if website:
        from urllib.parse import urlparse

        parsed = urlparse(website if "://" in website else f"https://{website}")
        domain = parsed.netloc.lower().replace("www.", "")

    company = Company(
        name=company_name.strip(),
        website=website.strip() or None,
        domain=domain or None,
        phone=phone.strip() or None,
        hq_city=city.strip() or None,
        hq_state=state.strip() or None,
        hq_country=country.strip() or "US",
        source="ai_lookup",
        is_active=True,
    )
    db.add(company)
    db.flush()

    # Create default site
    site_name = city.strip() or "HQ"
    site = CustomerSite(
        company_id=company.id,
        site_name=site_name,
        address_line1=address_line1.strip() or None,
        city=city.strip() or None,
        state=state.strip() or None,
        zip=zip.strip() or None,
        country=country.strip() or "US",
        contact_phone=phone.strip() or None,
    )
    db.add(site)
    db.commit()

    invalidate_prefix("company_list")

    display = html_mod.escape(f"{company.name} — {site.site_name}")

    return HTMLResponse(
        f'<div class="mt-1 p-2 bg-emerald-50 border border-emerald-200 rounded text-xs text-emerald-700">'
        f"Created: {display}"
        f"</div>"
        f'<span class="hidden" data-site-id="{site.id}" data-display="{display}"></span>'
    )
